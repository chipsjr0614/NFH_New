#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_nfh.py — 칩스가 편집한 엑셀(간호기록_원본_*.xlsx)을 읽어
app/nfh_data.js (window.NFH_RAW)로 빌드하고, 진술문을 마스터와 대조한다.

사용:  python3 data/build_nfh.py [엑셀경로]
       (생략 시 data/ 안의 가장 최근 간호기록_원본_*.xlsx 사용)

데이터 무결성: 진술문이 마스터(Nr.statement_*.xlsx)를 벗어나면 '경고'만 출력하고
빌드는 진행한다(막지 않음). 값은 임의로 고치지 않는다.
"""
import json, os, sys, glob, re
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SVC  = os.path.dirname(HERE)
APP  = os.path.join(SVC, "app")
MASTER = os.path.join(SVC, "Nr.statement_1023_Final.xlsx")

def find_src():
    if len(sys.argv)>1: return sys.argv[1]
    cands = sorted(glob.glob(os.path.join(HERE, "간호기록_원본_*.xlsx")))
    if not cands: sys.exit("엑셀 원본을 찾을 수 없습니다 (data/간호기록_원본_*.xlsx)")
    return cands[-1]

SRC = find_src()
wb = openpyxl.load_workbook(SRC, data_only=True)

def rows(sheet):
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r): continue
        yield {h:(v if v is not None else "") for h,v in zip(headers,r)}

# ── 사정세트 → DATA ──
DATA = {}
PROCMAP = {"A&E":"사정 및 평가(A&E)","D":"간호진단(D)","P&E":"계획 및 중재(P&E)"}
acc = {}  # (대,소) -> {full: [item,...]}
order = []
for r in rows("2.사정세트"):
    dae, dso, proc = str(r["대분류"]).strip(), str(r["소분류"]).strip(), str(r["과정"]).strip()
    txt = str(r["진술문(마스터원문)"]).strip()
    if not (dae and dso and proc and txt): continue
    key=(dae,dso)
    if key not in acc: acc[key]={"사정 및 평가(A&E)":[],"간호진단(D)":[],"계획 및 중재(P&E)":[]}; order.append(key)
    item = {"text":txt, "note":str(r.get("비고","")).strip()}
    if proc=="A&E":
        q=str(r.get("질문문구(💬)","")).strip()
        fmt=str(r.get("입력형식","")).strip()
        req=str(r.get("필수","")).strip().upper()
        if q:   item["q"]=q
        if fmt: item["format"]=fmt
        if req=="Y": item["required"]=True
    acc[key][PROCMAP[proc]].append(item)
for (dae,dso) in order:
    DATA.setdefault(dae,{})[dso] = {"사정": acc[(dae,dso)]}

# ── 주호소트리 → SYMS / BRANCHES ──
SYMS=[]; BRANCHES={}
tree={}  # symid -> {label,icon,rows:[]}
sym_order=[]
for r in rows("1.주호소트리"):
    sid=str(r["주증상id"]).strip()
    if not sid: continue
    if sid not in tree:
        tree[sid]={"label":str(r["주증상"]).strip(),"icon":str(r.get("아이콘","")).strip(),"rows":[]}
        sym_order.append(sid)
    tree[sid]["rows"].append(r)

def cell(r,k): return str(r.get(k,"")).strip()
for sid in sym_order:
    t=tree[sid]; sym={"id":sid,"label":t["label"],"icon":t["icon"]}
    rws=t["rows"]
    has_q1=any(cell(r,"질문1") for r in rws)
    if not has_q1:
        r0=rws[0]; dae,dso=cell(r0,"대분류"),cell(r0,"소분류")
        if dae and dso: sym["direct"]=[dae,dso]
        else:           sym["skipToFall"]=True
    else:
        bid=sid; sym["branch"]=bid
        q1=next(cell(r,"질문1") for r in rws if cell(r,"질문1"))
        opts=[]; seen1=[]
        for r in rws:
            s1=cell(r,"선택1")
            if s1 and s1 not in seen1: seen1.append(s1)
        for idx,s1 in enumerate(seen1):
            sub=[r for r in rws if cell(r,"선택1")==s1]
            if any(cell(r,"질문2") for r in sub):     # 2차 분기
                subid=f"{sid}__{idx}"
                q2=next(cell(r,"질문2") for r in sub if cell(r,"질문2"))
                o2=[{"label":cell(r,"선택2"),"key":[cell(r,"대분류"),cell(r,"소분류")]}
                    for r in sub if cell(r,"대분류")]
                BRANCHES[subid]={"q":q2,"hint":q2,"opts":o2}
                opts.append({"label":s1,"next":subid})
            else:
                r=sub[0]
                opts.append({"label":s1,"key":[cell(r,"대분류"),cell(r,"소분류")]})
        BRANCHES[bid]={"q":q1,"hint":q1,"opts":opts}
    SYMS.append(sym)

# ── 환자설명 / 도관 / 낙상교육 ──
EXPLAIN_SCRIPTS={}
for r in rows("3.환자설명"):
    dx=str(r["진단"]).strip()
    if not dx: continue
    EXPLAIN_SCRIPTS.setdefault(dx,[]).append({"say":str(r["말할 문구(📢)"]).strip(),"pe":str(r["연동 P&E"]).strip()})
CATS=[str(r["도관 진술문"]).strip() for r in rows("4.도관") if str(r["도관 진술문"]).strip()]
EDU_ITEMS=[]
for r in rows("5.낙상교육"):
    rid=str(r["id"]).strip()
    if not rid: continue
    recs=[x.strip() for x in str(r["기록 진술문(여러개는 | 로 구분)"]).split("|") if x.strip()]
    EDU_ITEMS.append({"id":rid,"say":str(r["말할 문구(📢)"]).strip(),"records":recs})

# ── 마스터 대조 (경고만) ──
def load_master():
    if not os.path.exists(MASTER): return None
    mwb=openpyxl.load_workbook(MASTER, read_only=True, data_only=True)
    s=set()  # 마스터 전체 시트의 모든 텍스트 셀 (진술문이 여러 시트에 분산되어 있어 폭넓게 수집)
    for ws in mwb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v,str) and v.strip(): s.add(v.strip())
    return s

def parts_of(text):
    """슬래시형을 완전한 진술문으로 확장해 대조용 조각 목록을 만든다.
    '복부 불편감 있음/없음' → ['복부 불편감 있음','복부 불편감 없음']
    (첫 항목을 기준 문장으로 보고, 뒤의 짧은 꼬리에 접두어를 잇는다)"""
    t=text.strip()
    if t.endswith(":"): return []                  # 값입력형 — 정형 진술문 아님(대조 제외)
    if "/" not in t: return [t]
    raw=[p.strip() for p in t.split("/") if p.strip()]
    if not raw: return []
    base=raw[0]; toks=base.split(" ")
    prefix=" ".join(toks[:-1]) if len(toks)>1 else ""
    out=[base]
    for p in raw[1:]:
        out.append(p if (" " in p or not prefix) else prefix+" "+p)
    return out

master=load_master()
warnings=[]
if master is not None:
    for dae,subs in DATA.items():
        for dso,blk in subs.items():
            for it in blk["사정"]["사정 및 평가(A&E)"]+blk["사정"]["계획 및 중재(P&E)"]:
                for p in parts_of(it["text"]):
                    if p and p not in master:
                        warnings.append(f"[{dae}/{dso}] {it['text']}  (조각 '{p}' 마스터에 없음)")

# ── 출력 ──
RAW={"DATA":DATA,"SYMS":SYMS,"BRANCHES":BRANCHES,
     "EXPLAIN_SCRIPTS":EXPLAIN_SCRIPTS,"EDU_ITEMS":EDU_ITEMS,"CATS":CATS}
js = ("// nfh_data.js — build_nfh.py 자동생성 (직접 수정 금지)\n"
      "// 원본: data/" + os.path.basename(SRC) + "\n"
      "window.NFH_RAW = " + json.dumps(RAW, ensure_ascii=False) + ";\n")
out = os.path.join(APP,"nfh_data.js")
open(out,"w",encoding="utf-8").write(js)

if master is not None and warnings:
    rep=os.path.join(HERE,"_마스터미일치.txt")
    with open(rep,"w",encoding="utf-8") as fp:
        fp.write(f"마스터 미일치 {len(warnings)}건 — 칩스 확인 (진술문이 마스터에 그 표기로 없음)\n")
        fp.write("="*60+"\n")
        for w in warnings: fp.write(w+"\n")

print("빌드 완료 →", out)
print(f"  DATA {sum(len(v) for v in DATA.values())} 소분류 / SYMS {len(SYMS)} / BRANCHES {len(BRANCHES)} / 환자설명 {len(EXPLAIN_SCRIPTS)} / 도관 {len(CATS)} / 낙상교육 {len(EDU_ITEMS)}")
if master is None:
    print("  ⚠ 마스터 파일 없음 — 진술문 대조 건너뜀")
elif warnings:
    print(f"  ⚠ 마스터 미일치 {len(warnings)}건 (값입력형/표기변형 다수 — 칩스 확인):")
    for w in warnings[:15]: print("     -", w)
    if len(warnings)>15: print(f"     … 외 {len(warnings)-15}건")
else:
    print("  ✓ 모든 진술문 마스터 일치")

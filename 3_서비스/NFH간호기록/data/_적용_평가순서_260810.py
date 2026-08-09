"""평가 순서 손보기 — 이제 기록에 나갈 2개를 「엑셀 순서」가 정하므로,
중요한 것이 앞에 오도록 몇 군데를 바로잡는다.

기존 순서 값 자체는 그대로 두고 「그 안에서의 자리」만 바꾼다.
그래야 같은 경로의 다른 진단 평가와 섞인 번호가 안 흐트러진다.
"""
import openpyxl, shutil, datetime, os, collections

BASE = '/Users/sungkwanchoi/Library/Mobile Documents/com~apple~CloudDocs/AI/NFH_New/3_서비스/NFH간호기록/data'
SRC = os.path.join(BASE, '간호기록_원본_260623.xlsx')
BK = os.path.join(BASE, '_backup', '간호기록_원본_260623_평가순서전_%s.xlsx'
                  % datetime.datetime.now().strftime('%y%m%d_%H%M%S'))
shutil.copy2(SRC, BK)
print('백업:', os.path.basename(BK))

wb = openpyxl.load_workbook(SRC)
ws = wb['2.사정세트']
H = [c.value for c in ws[1]]
c소, c과, c순, c문 = (H.index(x) + 1 for x in
                     ('소분류', '과정', '순서', '진술문(마스터원문)'))

# 「이 묶음은 이 차례로」 — 앞에 적힌 것이 기록에 먼저 나간다
GROUPS = [
    # 통증: 강도 → 인정 → 양상 → 관찰(계획이라 뒤로)
    ['통증 강도 측정함', '통증을 인정해줌', '통증 양상을 확인함', '통증을 관찰하기로 함'],
    # 저체온: 덮어주는 게 먼저다
    ['담요를 덮어줌', '체온 측정함', '저체온의 초기 증상에 대해 설명함', '사지 근력 평가함'],
]

# 소분류 → {진술문: [행번호]}
byset = collections.defaultdict(lambda: collections.defaultdict(list))
for r in range(2, ws.max_row + 1):
    if str(ws.cell(r, c과).value or '').strip() != 'P&E':
        continue
    소 = str(ws.cell(r, c소).value or '').strip()
    t = str(ws.cell(r, c문).value or '').strip()
    if 소 and t:
        byset[소][t].append(r)

changed = 0
for 소, m in byset.items():
    for order in GROUPS:
        rows = [(t, r) for t in order for r in m.get(t, [])]
        if len(rows) < 2:
            continue
        vals = sorted(ws.cell(r, c순).value for _, r in rows)
        for (t, r), v in zip(rows, vals):
            if ws.cell(r, c순).value != v:
                ws.cell(r, c순).value = v
                changed += 1
                print('  %-28s %-26s → 순서 %s' % (소[:28], t[:26], v))

wb.save(SRC)
print('\n%d칸 조정 · 저장 완료' % changed)

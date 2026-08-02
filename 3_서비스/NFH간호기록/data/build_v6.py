#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
간호기록 V6 데이터 빌드
  간호기록_원본_*.xlsx  →  질문은행 + 경로  →  nfh_v6_data.js

핵심 원칙 (CLAUDE.md 절대규칙)
  · 진술문은 진술문 마스터 원문만 사용한다.
  · 자동 교정은 「띄어쓰기·대소문자만 다르고 마스터에 원문이 확실히 있는 경우」로 한정한다.
  · 마스터에 없는 문구는 지어내지 않고 _확인필요.txt 에 보고만 한다.

실행:  python3 build_v6.py
"""
import os, re, json, fnmatch, collections, unicodedata
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SVC  = os.path.dirname(HERE)                      # 3_서비스/NFH간호기록
OUT_JS   = os.path.join(SVC, 'app', 'nfh_v6_data.js')
OUT_WARN = os.path.join(HERE, '_확인필요.txt')
TEMPLATE = os.path.join(HERE, 'v6_template.html')
OUT_HTML = os.path.abspath(os.path.join(SVC, '..', '..', 'nursing', 'app', '간호기록V6.html'))

# ── 입력 파일 ────────────────────────────────────────────────
# macOS 파일명은 자소분리(NFD)로 저장되는데 이 소스의 한글은 NFC라
# glob 패턴이 그냥은 안 맞는다. 양쪽을 NFC로 맞춰놓고 비교한다.
nfc = lambda s: unicodedata.normalize('NFC', s)

def newest(pattern):
    pat = nfc(pattern)
    c = sorted((f for f in os.listdir(HERE)
                if fnmatch.fnmatch(nfc(f), pat) and '백업' not in nfc(f)), key=nfc)
    if not c: raise SystemExit(f'입력 파일 없음: {pattern}')
    return os.path.join(HERE, c[-1])

SRC    = newest('간호기록_원본_*.xlsx')
MASTER = os.path.join(SVC, 'Nr.statement_1023_Final.xlsx')

# ── 마스터 적재 ──────────────────────────────────────────────
nsp  = lambda s: re.sub(r'\s', '', str(s))          # 공백 무시
nspc = lambda s: nsp(s).lower()                     # 공백+대소문자 무시

wbm = openpyxl.load_workbook(MASTER, read_only=True)
M_SP, M_SPC = {}, {}
DX_SET = set()
for r in wbm['진술문'].iter_rows(min_row=1, values_only=True):
    if not r[1]: continue
    n, s = r[0], str(r[1]).strip()
    M_SP.setdefault(nsp(s), s)
    M_SPC.setdefault(nspc(s), s)
    if isinstance(n, int) and 7875 <= n <= 8079:
        DX_SET.add(nspc(s))

WARN = []
def canon(text, where):
    """마스터 원문으로 정규화. 없으면 원문 그대로 두고 경고."""
    t = str(text).strip()
    if nsp(t) in M_SP:
        fixed = M_SP[nsp(t)]
        if fixed != t: WARN.append(('띄어쓰기교정', where, t, fixed))
        return fixed
    if nspc(t) in M_SPC:
        fixed = M_SPC[nspc(t)]
        WARN.append(('대소문자교정', where, t, fixed))
        return fixed
    WARN.append(('마스터없음', where, t, ''))
    return t

# ── 토글/선택 진술문 분해 ─────────────────────────────────────
SHORT = ('있음','없음','안됨','됨','강함','약함','가능함','가능하지 않음','정상임','멈춤','좋음')
def split_stmt(v):
    """'열감 있음/없음' → ['열감 있음','열감 없음']  (축약형 확장)"""
    parts = [x.strip() for x in str(v).split('/')]
    if len(parts) == 1: return parts
    out = [parts[0]]
    for x in parts[1:]:
        if x in SHORT:
            stem = parts[0]
            for sh in sorted(SHORT, key=len, reverse=True):
                if parts[0].endswith(sh):
                    stem = parts[0][:-len(sh)].strip(); break
            out.append(f'{stem} {x}'.strip() if stem and stem != parts[0] else x)
        else:
            out.append(x)
    return out

def rows(sheet):
    ws = openpyxl.load_workbook(SRC, read_only=True, data_only=True)[sheet]
    it = ws.iter_rows(values_only=True)
    hdr = [str(h).strip() if h is not None else '' for h in next(it)]
    for r in it:
        d = {hdr[i]: r[i] for i in range(min(len(hdr), len(r)))}
        if any(v is not None and str(v).strip() for v in d.values()):
            yield d

# ── 1) 질문은행 + 경로 ───────────────────────────────────────
QBANK   = collections.OrderedDict()   # qid -> {q, fmt, stmts, core}
QINDEX  = {}                          # (질문문구, 진술문원본) -> qid
PATHSET = collections.OrderedDict()   # 소분류 -> {대분류, dx:[{name, ae:[qid], pe:[문구]}], ae_all:[qid]}

def qkey(q):
    """중복 판정용 질문 키. 앞머리 🩺/💬 표시는 화면 힌트일 뿐이라 무시한다."""
    return nsp(re.sub(r'^[^\w"“\'가-힣]+', '', str(q)))

def qid_for(q, raw, fmt, where, det=''):
    # 표기 변형을 흡수하려면 「마스터 원문으로 정규화한 뒤」 중복을 판단해야 한다.
    stmts = [canon(x, where) for x in split_stmt(raw)]
    key = (qkey(q), tuple(stmts), det)
    if key in QINDEX: return QINDEX[key]
    qid = f'Q{len(QBANK)+1:03d}'
    QBANK[qid] = {'q': str(q).strip(), 'fmt': str(fmt or '체크').strip(),
                  'stmts': stmts, 'paths': set(), 'det': det}
    QINDEX[key] = qid
    return qid

for r in rows('2.사정세트'):
    대, 소 = str(r.get('대분류') or '').strip(), str(r.get('소분류') or '').strip()
    과 = str(r.get('과정') or '').strip()
    raw = r.get('진술문(마스터원문)')
    if not 소 or not 과 or raw is None: continue
    link = str(r.get('진단연결') or '').strip()
    P = PATHSET.setdefault(소, {'대분류': 대, 'dx': collections.OrderedDict(), 'items': []})
    where = f'{소}/{과}'

    if 과 == 'D':
        name = canon(raw, where)
        if nspc(name) not in DX_SET:
            WARN.append(('진단아님', where, name, '마스터 간호진단 목록에 없음'))
        P['dx'].setdefault(name, {'ae': [], 'pe': []})
    elif 과 == 'A&E':
        qid = qid_for(r.get('질문문구(💬)') or '', raw, r.get('입력형식'), where,
                      str(r.get('상세') or '').strip())
        QBANK[qid]['paths'].add(소)
        P['items'].append(('ae', qid, link))
    elif 과 == 'P&E':
        P['items'].append(('pe', canon(raw, where), link))

# 진단별 배치 (진단연결 사용, '공통'은 전 진단에)
# 진단명은 canon으로 마스터 원문이 되는데(「비효율적 호흡양상」→「비효율적호흡양상」)
# 진단연결 열은 엑셀 표기 그대로다. 글자로 비교하면 안 맞아 전부 「공통」으로 새어
# 같은 줄이 진단마다 반복된다. 공백·대소문자를 무시하고 맞춘다.
for 소, P in PATHSET.items():
    names = list(P['dx'].keys())
    nmap = {nspc(n): n for n in names}
    for kind, val, link in P['items']:
        hit = nmap.get(nspc(link)) if link else None
        tgt = names if (link == '공통' or hit is None) else [hit]
        if not names: continue
        for n in tgt: P['dx'][n][kind].append(val)

# ── 2.5) 통증 공통 (7.통증공통) ──────────────────────────────
# 통증 위젯이 쓰는 문항·값을 코드가 아니라 엑셀에서 읽는다.
# 진통제 문항을 21개 경로에 중복으로 적지 않도록, 여기 한 번만 적고 빌드가 붙인다.
PAIN = {'ae': [], 'grp': {}, 'aeCond': [], 'peCond': [], 'parts': [], 'subs': {},
        'tools': [], 'patterns': [], 'freqs': [], 'durs': [], 'chronic': '',
        'smc': [], 'smcDx': [], 'dxAcute': '', 'dxChronic': '', 'months': 3}
for r in rows('7.통증공통'):
    과 = str(r.get('과정') or '').strip()
    구 = str(r.get('구분') or '').strip()
    값 = str(r.get('값') or '').strip()
    비 = str(r.get('비고') or '').strip()
    그 = str(r.get('그룹') or '').strip()
    if 과 == 'A&E':
        qid = qid_for(r.get('질문문구(💬)') or '', r.get('진술문(마스터원문)'),
                      r.get('입력형식'), '통증공통/A&E')
        PAIN['ae'].append(qid)
        if 그: PAIN['grp'][qid] = 그
    elif 과 == 'A&E조건':
        PAIN['aeCond'].append({'cond': str(r.get('조건') or '').strip(),
                               'rec': canon(r.get('진술문(마스터원문)'), '통증공통/A&E조건'),
                               'q':   str(r.get('질문문구(💬)') or '').strip(),
                               'fmt': str(r.get('입력형식') or '').strip()})
    elif 과 == 'P&E조건':
        PAIN['peCond'].append({'cond': str(r.get('조건') or '').strip(),
                               'rec': canon(r.get('진술문(마스터원문)'), '통증공통/P&E')})
    elif 과 == '값':
        if   구 == '기간':
            PAIN['durs'].append(값)
            if 그 == '만성': PAIN['chronic'] = 값
        elif 구 == '부위':     PAIN['parts'].append({'v': 값, 'ic': 비, 'smc': 그 == 'SMC'})
        elif 구 == '세부부위': PAIN['subs'].setdefault(비, []).append(값)
        elif 구 == '도구':     PAIN['tools'].append({'v': 값, 'd': 비, 'ped': 그 == '소아'})
        elif 구 == '양상':     PAIN['patterns'].append(값)
        elif 구 == '빈도':     PAIN['freqs'].append(값)
        elif 구 == 'SMC':      PAIN['smc'].append(값)
        elif 구 == 'SMC진단':  PAIN['smcDx'].append(값)
        elif 구 == '기간기준':
            try: PAIN['months'] = int(float(값))
            except ValueError: WARN.append(('마스터없음', '통증공통/기간기준', 값, ''))
        elif 구 == '통증진단':
            name = canon(값, '통증공통/진단')
            if nspc(name) not in DX_SET:
                WARN.append(('진단아님', '통증공통/진단', name, '마스터 간호진단 목록에 없음'))
            if str(r.get('순서') or '1').strip() == '1': PAIN['dxAcute'] = name
            else: PAIN['dxChronic'] = name

# 그룹이 없는 문항(진통제 원함)만 경로 질문으로 붙인다.
# 「투여후」·「SMC」 그룹은 통증 위젯 안에서만 쓰이므로 경로 질문에 넣지 않는다.
if PAIN['dxAcute']:
    for 소, P in PATHSET.items():
        d = P['dx'].get(PAIN['dxAcute'])
        if d is None: continue
        for qid in PAIN['ae']:
            if PAIN['grp'].get(qid): continue
            if qid not in d['ae']: d['ae'].append(qid)
            QBANK[qid]['paths'].add(소)

# ── 2.6) 신경학 공통 (8.신경학공통) ─────────────────────────
# 기존 경로의 신경학 행은 건드리지 않는다. 문구가 경로마다 다르고 진단연결도 달라서
# 공통으로 합치면 임상 내용이 조용히 바뀐다. 여기 것은 「항목이 없는 경로」에만 붙인다.
NEURO = []
for r in rows('8.신경학공통'):
    raw = r.get('진술문(마스터원문)')
    if raw is None: continue
    NEURO.append(qid_for(r.get('질문문구(💬)') or '', raw, r.get('입력형식'), '신경학공통'))

sig = lambda qid: frozenset(nspc(s) for s in QBANK[qid]['stmts'])
NEURO_SIG = {qid: sig(qid) for qid in NEURO}
PATH_NEURO = {}
for 소, P in PATHSET.items():
    have = {sig(val) for kind, val, _link in P['items'] if kind == 'ae'}
    PATH_NEURO[소] = [q for q in NEURO if NEURO_SIG[q] not in have]

# ── 2.7) 추천 근거 (9.추천근거) ──────────────────────────────
# 「있음이 이상」이라는 규칙은 기계로 못 만든다. 구토는 있는 게 이상이고
# 지남력은 없는 게 이상이라 글자만 봐서는 갈리지 않는다. 칩스가 지정한 것만 쓴다.
ABNORMAL, ABN_TEXT = set(), set()
for r in rows('9.추천근거'):
    s = str(r.get('이상소견 진술문') or '').strip()
    if not s: continue
    t = canon(s, '추천근거')
    ABNORMAL.add(nspc(t)); ABN_TEXT.add(t)
for qid, v in QBANK.items():
    v['ab'] = 1 if (v['fmt'] == '토글' and len(v['stmts']) >= 2
                    and nspc(v['stmts'][1]) in ABNORMAL) else 0

# ── 2.75) 자동 평가 (11.평가자동) ───────────────────────────
# 「출혈양상 확인함」은 출혈 사정을 했으면 이미 한 것이다. 또 누르게 할 이유가 없다.
# 여기 적힌 것은 그 진단의 사정을 하나라도 답하면 자동으로 켜진다 (끌 수는 있다).
AUTO_PE = []
for r in rows('11.평가자동'):
    t = str(r.get('진술문(마스터원문)') or '').strip()
    if not t or t.startswith('—') or t.startswith('(수동)'): continue
    AUTO_PE.append(canon(t, '평가자동'))
AUTO_PE = list(dict.fromkeys(AUTO_PE))

# ── 2.78) 부위·상황 (12.부위상세) ───────────────────────────
# 부위마다 경로를 만들면 끝이 없다. 경로는 하나로 두고 부위·상황을 골라
# 진술문 뒤에 괄호로 붙인다 — 「신체손상 있음(머리, 가슴)」.
# 고르면 그 부위의 문항과 진단이 따라 나온다. 구분(외상부위·피부상황)으로 묶는다.
SITES = collections.OrderedDict()
for r in rows('12.부위상세'):
    grp = str(r.get('구분') or '').strip()
    nm  = str(r.get('값') or '').strip()
    if not grp or not nm: continue
    G = SITES.setdefault(grp, collections.OrderedDict())
    S_ = G.setdefault(nm, {'v': nm, 'ic': str(r.get('아이콘') or '').strip(),
                           'dx': str(r.get('진단연결') or '').strip(),
                           'ae': [], 'pe': [], 'note': str(r.get('비고') or '').strip()})
    raw = r.get('진술문(마스터원문)')
    if raw is None or not str(raw).strip(): continue
    if str(r.get('과정') or '').strip() == 'A&E':
        S_['ae'].append(qid_for(r.get('질문문구(💬)') or '', raw, r.get('입력형식'), f'{grp}/{nm}'))
    else:
        S_['pe'].append(canon(raw, f'{grp}/{nm}'))

# ── 2.8) 후속 사정 (10.후속사정) ────────────────────────────
# 「신경학적 이상이 보이면 혈당을 재라」처럼, 앞의 답에 따라 뒤에 붙는 문항.
# 경고 배지가 아니라 「질문이 하나 늘어나는 것」이라 미응답 장치가 그대로 먹는다.
FOLLOW = collections.OrderedDict()
for r in rows('10.후속사정'):
    trig = str(r.get('발동조건') or '').strip()
    raw  = r.get('진술문(마스터원문)')
    if not trig or raw is None: continue
    F = FOLLOW.setdefault(trig, {'trig': trig, 'skip': str(r.get('생략조건') or '').strip(),
                                 'ae': [], 'pe': [], 'skipPaths': []})
    if str(r.get('과정') or '').strip() == 'A&E':
        F['ae'].append(qid_for(r.get('질문문구(💬)') or '', raw, r.get('입력형식'), '후속사정'))
    else:
        F['pe'].append(canon(raw, '후속사정/P&E'))
# 이미 그것을 묻고 있는 경로에서는 뜨지 않는다 (저혈당·고혈당 등)
for F in FOLLOW.values():
    if not F['skip']: continue
    for 소, P in PATHSET.items():
        if any(F['skip'] in s for kind, val, _l in P['items'] if kind == 'ae'
               for s in QBANK[val]['stmts']):
            F['skipPaths'].append(소)

# 신경학 문항 전체 — 경로에 원래 있던 것 + 버튼으로 붙는 것.
# 발동 판정은 둘 다 봐야 한다 (어지럼처럼 이미 다 있는 경로도 있다).
PATH_NEURO_ALL = {}
for 소, P in PATHSET.items():
    inline = [val for kind, val, _l in P['items']
              if kind == 'ae' and sig(val) in set(NEURO_SIG.values())]
    PATH_NEURO_ALL[소] = list(dict.fromkeys(inline + PATH_NEURO[소]))

# 같은 진술문이 두 번 들어가면 기록에 중복으로 찍힌다
for P in PATHSET.values():
    for d in P['dx'].values():
        d['ae'] = list(dict.fromkeys(d['ae']))
        d['pe'] = list(dict.fromkeys(d['pe']))

# ── 2) 공통 질문(core) 판정 — 경로 절반 이상 등장 ──────────────
NPATH = len(PATHSET)
for qid, v in QBANK.items():
    v['core'] = len(v['paths']) >= NPATH * 0.25
    v['paths'] = sorted(v['paths'])

# ── 3) 주호소 트리 ───────────────────────────────────────────
SYMS, BRANCH = collections.OrderedDict(), []
for r in rows('1.주호소트리'):
    sid = str(r.get('주증상id') or '').strip()
    nm  = str(r.get('주증상') or '').strip()
    if not nm: continue
    SYMS.setdefault(nm, {'id': sid, 'icon': str(r.get('아이콘') or '').strip(),
                         'q1': str(r.get('질문1') or '').strip(),
                         'q2': str(r.get('질문2') or '').strip()})
    BRANCH.append({'sym': nm,
                   'sel1': str(r.get('선택1') or '').strip(),
                   'sel2': str(r.get('선택2') or '').strip(),
                   'set':  str(r.get('소분류') or '').strip()})

# ── 4) 도관 / 낙상교육 / 환자설명 / 마무리질문 ─────────────────
CATH = []
for r in rows('4.도관'):
    raw = r.get('도관 진술문(| 로 구분)') or r.get('도관 진술문')
    if not r.get('도관') or not raw: continue
    CATH.append({'s': str(r['도관']).strip(),
                 'n': str(r.get('정식명칭') or r['도관']).strip(),
                 'd': [canon(x, f"도관/{r['도관']}") for x in str(raw).split('|') if x.strip()],
                 'new': bool(str(r.get('비고') or '').strip()),
                 'top': int(r.get('순서') or 99) <= 10})
EDU  = [{'say': str(r.get('말할 문구(📢)') or '').strip(),
         'recs': list(dict.fromkeys(canon(x, '낙상교육') for x in str(r.get('기록 진술문(여러개는 | 로 구분)') or '').split('|') if x.strip()))}
        for r in rows('5.낙상교육') if r.get('id')]
# 진단명은 공백 차이가 있어 정규화해 묶는다. P&E도 마스터 원문으로 맞춘다.
EXPL_RAW = collections.defaultdict(list)
for r in rows('3.환자설명'):
    if r.get('진단'):
        EXPL_RAW[nsp(r['진단'])].append({'say': str(r.get('말할 문구(📢)') or '').strip().strip('"'),
                                         'pe':  canon(r.get('연동 P&E') or '', '환자설명')})
CLOSING = [{'sel': str(r.get('선택') or '').strip(),
            'q':   str(r.get('질문문구(💬)') or '').strip(),
            'rec': canon(r['진술문(마스터원문)'], '마무리질문') if r.get('진술문(마스터원문)') else '',
            'next':str(r.get('다음동작') or '').strip()}
           for r in rows('6.마무리질문') if r.get('선택')]

# ── 5) 출력 ─────────────────────────────────────────────────
DATA = {
  'meta': {'src': os.path.basename(SRC), 'paths': NPATH, 'questions': len(QBANK)},
  'qbank': {k: {'q': v['q'], 'fmt': v['fmt'], 'stmts': v['stmts'], 'ab': v.get('ab', 0),
                'det': v.get('det', ''), 'core': v['core'], 'n': len(v['paths'])}
            for k, v in QBANK.items()},
  'syms': SYMS, 'branch': BRANCH,
  'sets': {k: {'대분류': v['대분류'], 'neuro': PATH_NEURO.get(k, []),
               'neuroAll': PATH_NEURO_ALL.get(k, []),
               'dx': [{'name': n, 'ae': d['ae'], 'pe': d['pe']} for n, d in v['dx'].items()]}
           for k, v in PATHSET.items()},
  # 이상소견 진술문 원문 — 선택형(빛 반사 등)은 이걸로 이상 여부를 가린다
  'sites': {g: list(v.values()) for g, v in SITES.items()},
  'abnormal': sorted({M_SP.get(a, a) for a in ABN_TEXT}),
  'follow': list(FOLLOW.values()),
  'autoPE': AUTO_PE,
  'cath': CATH, 'edu': EDU, 'closing': CLOSING, 'pain': PAIN,
  # 경로에서 실제 쓰는 진단명 기준으로 환자설명을 붙인다
  'explain': {n: list({e['pe']: e for e in EXPL_RAW.get(nsp(n), [])}.values())
              for n in {dx for P in PATHSET.values() for dx in P['dx']}},
}
os.makedirs(os.path.dirname(OUT_JS), exist_ok=True)
with open(OUT_JS, 'w', encoding='utf-8') as f:
    f.write('/* 자동 생성 — 직접 수정 금지. 데이터는 엑셀을 고치고 build_v6.py 실행 */\n')
    f.write('window.NFH_V6 = ' + json.dumps(DATA, ensure_ascii=False, indent=1) + ';\n')

# ── 6) 경고 리포트 (같은 문구는 1건으로 집계) ─────────────────
UNIQ = collections.OrderedDict()
for kind, where, a, b in WARN:
    UNIQ.setdefault((kind, a, b), where)
cnt = collections.Counter(k[0] for k in UNIQ)

# 같은 질문인데 진술문이 달라 갈린 것 → 선택형으로 합칠 후보
SPLIT = collections.defaultdict(list)
for qid, v in QBANK.items(): SPLIT[qkey(v['q'])].append(qid)
SPLIT = {k: v for k, v in SPLIT.items() if len(v) > 1}
VAGUE = [qid for qid, v in QBANK.items() if len(nsp(v['q'])) <= 6]

# 토글 극성 — 첫 진술문이 음성인 행. 기록 자체는 앱이 진술문 글씨를 버튼에
# 그대로 박아 안전하지만, 「진단 추천 근거」는 여전히 첫 진술문이라 확인이 필요하다.
NEG_END = ('없음', '안됨', '않음', '못함', '못 봄', '못봄', '불가함')
POS_END = ('있음', '됨', '함', '봄')
ends = lambda s, tup: any(s.endswith(x) for x in tup)
# 9.추천근거에 적힌 것은 이미 칩스가 본 것이므로 경고에서 뺀다
reviewed = lambda v: any(nspc(s) in ABNORMAL for s in v['stmts'])
POLAR = [qid for qid, v in QBANK.items()
         if v['fmt'] == '토글' and len(v['stmts']) >= 2 and not reviewed(v)
         and ends(v['stmts'][0], NEG_END)
         and ends(v['stmts'][1], POS_END) and not ends(v['stmts'][1], NEG_END)]
# ── 어휘 표준 검사 ─────────────────────────────────────────
# 같은 뜻인데 경로마다 다른 표현을 쓰면 EMR 어휘가 흩어진다.
# 「N경로가 쓰는 표준 표현이 이미 있다」를 알려 자주 쓰는 쪽으로 모이게 한다.
import itertools as _it
_core = lambda t: re.sub(r'\s*(있음|없음|함|됨|보임|호소함|상태임)$', '', t).strip()
_key  = lambda t: re.sub(r'[\s,·]', '', _core(t))
_neg  = lambda t: re.search(r'(없음|안됨)$', t)

USE = collections.Counter()
PATH_POS = {}
for 소, P in PATHSET.items():
    seen = set()
    for kind, val, _l in P['items']:
        if kind != 'ae': continue
        for st in QBANK[val]['stmts']:
            if not _neg(st): seen.add(st)
    PATH_POS[소] = list(seen)
    for st in seen: USE[st] += 1

_grp = collections.defaultdict(set)
for t in USE: _grp[_key(t)].add(t)

STD = []      # 드문 표현 → 표준 후보
for 소, ss in PATH_POS.items():
    for t in ss:
        if USE[t] > 2: continue
        for kk, alts in _grp.items():
            if kk == _key(t) or len(kk) < 2: continue
            if kk in _key(t) or _key(t) in kk:
                best = max(alts, key=lambda x: USE[x])
                if USE[best] >= 5: STD.append((소, t, USE[t], best, USE[best]))
                break

SAME = []     # 한 경로 안에서 뜻이 겹치는 짝
for 소, ss in PATH_POS.items():
    for a, b in _it.combinations(sorted(ss), 2):
        ka, kb = _key(a), _key(b)
        if len(ka) < 2 or len(kb) < 2 or ka == kb: continue
        if ka in kb or kb in ka: SAME.append((소, a, USE[a], b, USE[b]))

# 전체 감사 목록 — 어느 답이 「추천」을 띄우는지 한눈에 보이게
AUDIT = [(qid, v) for qid, v in QBANK.items() if v['fmt'] == '토글' and len(v['stmts']) >= 2]

with open(OUT_WARN, 'w', encoding='utf-8') as f:
    f.write(f'간호기록 V6 빌드 리포트\n원본: {os.path.basename(SRC)}\n')
    f.write('='*70 + '\n')
    for k, n in cnt.most_common(): f.write(f'  {k:12} {n:4}건\n')
    f.write(f'  {"질문문구없음":12} {len(VAGUE):4}건\n')
    f.write(f'  {"선택형합칠후보":12} {len(SPLIT):4}건\n')
    f.write(f'  {"토글극성확인":12} {len(POLAR):4}건\n')
    f.write('='*70 + '\n')
    f.write('\n※ 자동 교정(띄어쓰기·대소문자)은 빌드가 이미 마스터 원문으로 맞췄다. 엑셀은 고치지 않아도 된다.\n')
    f.write('※ 「마스터없음」은 AI가 지어내지 않고 그대로 두었다. 칩스 판단이 필요하다.\n')

    for kind in ['마스터없음', '진단아님', '대소문자교정', '띄어쓰기교정']:
        sel = [(k, w) for k, w in UNIQ.items() if k[0] == kind]
        if not sel: continue
        f.write(f'\n\n[{kind}] {len(sel)}건\n' + '-'*70 + '\n')
        for (_, a, b), where in sel:
            f.write(f'  {where:32} 「{a}」' + (f'  →  「{b}」' if b else '') + '\n')

    if VAGUE:
        f.write(f'\n\n[질문문구없음] {len(VAGUE)}건 — 간호사가 뭘 물어야 할지 알 수 없음\n' + '-'*70 + '\n')
        for qid in VAGUE:
            f.write(f'  {qid}  질문「{QBANK[qid]["q"]}」  진술문 {QBANK[qid]["stmts"]}\n')

    if POLAR:
        f.write(f'\n\n[토글 극성 확인] {len(POLAR)}건 — 첫 진술문이 음성이다\n' + '-'*70 + '\n')
        f.write('기록은 안전하다. 앱이 버튼에 진술문 글씨를 그대로 박으므로 누른 대로 남는다.\n')
        f.write('확인이 필요한 것은 「진단 추천」이다. 추천은 아직 첫 진술문을 근거로 삼는다.\n')
        f.write('아래에서 이상소견(추천 근거가 되어야 할 쪽)이 어느 것인지 봐주세요.\n\n')
        for qid in POLAR:
            v = QBANK[qid]
            f.write(f'  {qid}  질문「{v["q"]}」  경로 {v["n"] if "n" in v else len(v["paths"])}개\n')
            f.write(f'        추천 근거(현재) : {v["stmts"][0]}\n')
            f.write(f'        반대쪽          : {v["stmts"][1]}\n')

    f.write(f'\n\n[어휘 표준] 자주 쓰는 표현 상위 12\n' + '-'*70 + '\n')
    f.write('여러 경로가 함께 쓰는 표현이다. 새 문항을 넣을 때 이 어휘를 먼저 쓴다.\n\n')
    for t, n in USE.most_common(12): f.write(f'  {n:>3}경로  {t}\n')

    if STD:
        f.write(f'\n\n[어휘 표준] 표준으로 모을 후보 {len(STD)}건\n' + '-'*70 + '\n')
        f.write('드문 표현을 쓰는데 같은 뜻으로 여러 경로가 쓰는 표준이 있다.\n')
        f.write('부위·정도가 임상적으로 다르면 그대로 두는 게 맞다. 판단이 필요하다.\n\n')
        for 소, t, n, best, bn in STD:
            f.write(f'  [{소[:22]}] 「{t}」({n}경로)  →  표준 「{best}」({bn}경로)\n')

    if SAME:
        f.write(f'\n\n[어휘 표준] 한 경로 안에서 뜻이 겹치는 짝 {len(SAME)}건\n' + '-'*70 + '\n')
        f.write('간호사가 비슷한 것을 두 번 묻는다. 정도·부위가 다르면 그대로 둔다.\n\n')
        for 소, a, na, b, nb in SAME:
            f.write(f'  [{소[:22]}] 「{a}」({na})  ↔  「{b}」({nb})\n')

    f.write(f'\n\n[추천 근거 감사] 토글 {len(AUDIT)}문항 — 어느 답이 「추천」을 띄우는가\n' + '-'*70 + '\n')
    f.write('★ 표시는 9.추천근거 시트에서 칩스가 지정한 것이다.\n')
    f.write('아래를 훑어보다 「이건 반대인데」 싶은 줄이 있으면 9.추천근거에 진술문을 적으면 된다.\n\n')
    for qid, v in AUDIT:
        mark = '★' if v.get('ab', 0) else ' '
        f.write(f'  {mark}{qid}  추천={v["stmts"][v.get("ab",0)]}\n')
        f.write(f'         반대={v["stmts"][1-v.get("ab",0)]}   질문「{v["q"][:44]}」\n')

    if SPLIT:
        f.write(f'\n\n[선택형 합칠 후보] {len(SPLIT)}건 — 같은 질문인데 진술문이 달라 따로 잡혔다\n' + '-'*70 + '\n')
        for _, qids in SPLIT.items():
            f.write(f'  질문「{QBANK[qids[0]]["q"]}」\n')
            for q in qids: f.write(f'      {q}  {QBANK[q]["stmts"]}\n')

# ── 단일 HTML 생성 (데이터 인라인) ──
if os.path.exists(TEMPLATE):
    tpl = open(TEMPLATE, encoding='utf-8').read()
    inline = 'window.NFH_V6 = ' + json.dumps(DATA, ensure_ascii=False, separators=(',', ':')) + ';'
    html = tpl.replace('/*__DATA__*/', inline)
    os.makedirs(os.path.dirname(OUT_HTML), exist_ok=True)
    with open(OUT_HTML, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f'✅ {OUT_HTML}  ({len(html)//1024} KB · 자체완결 단일 HTML)')

print(f'✅ {OUT_JS}')
print(f'   경로 {NPATH} · 질문 {len(QBANK)}개 (공통 {sum(1 for v in QBANK.values() if v["core"])}개)')
print(f'   도관 {len(CATH)} · 낙상교육 {len(EDU)} · 환자설명 {len(DATA["explain"])}종 · 마무리 {len(CLOSING)}')
print(f'📝 {OUT_WARN}')
for k, n in cnt.most_common(): print(f'   {k:12} {n:4}건')

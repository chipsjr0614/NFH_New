#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
seed_from_current.py — (일회성) 현재 앱 데이터(app/nfh_core.js + _archive/nursing_data.json)를
칩스 편집용 스타터 엑셀(간호기록_원본_YYMMDD.xlsx)로 역생성한다.

이후 관리는: 엑셀 편집 → build_nfh.py → app/nfh_data.js → 앱 반영.
한글 안전을 위해 Python(openpyxl)만 사용. (sed/perl 금지)
"""
import json, subprocess, os, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

HERE = os.path.dirname(os.path.abspath(__file__))
SVC  = os.path.dirname(HERE)
APP  = os.path.join(SVC, "app")
OUT  = os.path.join(HERE, "간호기록_원본_260621.xlsx")

# ── 1) node로 현재 NFH 데이터 + getQ 매핑 덤프 ──
NODE = r'''
global.window=global; require(process.argv[1]);
const N=window.NFH; const qmap={};
for(const c in N.DATA) for(const s in N.DATA[c]) for(const b in N.DATA[c][s]){
  (N.DATA[c][s][b]['사정 및 평가(A&E)']||[]).forEach(it=>{ if(!(it.text in qmap)) qmap[it.text]=N.getQ(it.text); });
}
process.stdout.write(JSON.stringify({SYMS:N.SYMS,BRANCHES:N.BRANCHES,EXPLAIN:N.EXPLAIN_SCRIPTS,EDU:N.EDU_ITEMS,CATS:N.CATS,DATA:N.DATA,qmap}));
'''
res = subprocess.run(["node","-e",NODE, os.path.join(APP,"nfh_core.js")],
                     capture_output=True, text=True)
if res.returncode!=0:
    print("node 덤프 실패:\n", res.stderr); sys.exit(1)
D = json.loads(res.stdout)
SYMS, BRANCHES, EXPLAIN, EDU, CATS, DATA, QMAP = (
    D["SYMS"], D["BRANCHES"], D["EXPLAIN"], D["EDU"], D["CATS"], D["DATA"], D["qmap"])

def detect_fmt(t):
    t=t.strip()
    if t.endswith(":"): return "값입력"
    if "/" in t:        return "토글"
    return "체크"

# ── 2) 엑셀 작성 ──
wb = openpyxl.Workbook()
HEAD = Font(bold=True, color="FFFFFF")
HFILL= PatternFill("solid", fgColor="2F5496")
WRAP = Alignment(vertical="top", wrap_text=True)

def sheet(title, headers, widths):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in ws[1]:
        c.font=HEAD; c.fill=HFILL; c.alignment=Alignment(vertical="center")
    for i,w in enumerate(widths,1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
    ws.freeze_panes="A2"
    return ws

# 시트0 설명
ws0 = wb.active; ws0.title="📖설명"
for row in [
 ["NFH 간호기록 — 데이터 원본 (칩스 편집용)"],
 [""],
 ["■ 편집 후: 터미널에서  python3 data/build_nfh.py  실행 → app/nfh_data.js 자동 생성 → 앱 반영"],
 ["■ '진술문' 열은 마스터(Nr.statement_*.xlsx) 원문만. build 시 자동 대조(벗어나면 경고)."],
 ["■ 빈 칸 규칙: 질문문구 비우면 자동 생성문 사용 / 입력형식 비우면 자동판별 / 필수는 Y만 표기."],
 [""],
 ["[시트 안내]"],
 ["1.주호소트리  : 주증상 → (분기질문/선택) → 사정세트(대분류·소분류) 연결. 질문 비면 분기없이 바로 사정."],
 ["2.사정세트    : 사정세트별 A&E/D/P&E 진술문. 질문문구·입력형식·필수(⚠️)·비고를 여기서 관리."],
 ["3.환자설명    : 진단별 '말할 문구' ↔ 연동 P&E."],
 ["4.도관        : 라인·튜브 체크리스트."],
 ["5.낙상교육    : 낙상 교육 멘트 ↔ 기록 진술문."],
]:
    ws0.append(row)
ws0.column_dimensions["A"].width=110

# 시트1 주호소트리
t = sheet("1.주호소트리",
    ["주증상id","주증상","아이콘","질문1","선택1","질문2","선택2","대분류","소분류","비고"],
    [12,18,8,26,18,26,18,30,40,20])
def add_tree(sym, q1="",s1="",q2="",s2="",dae="",dso="",memo=""):
    t.append([sym["id"],sym["label"],sym.get("icon",""),q1,s1,q2,s2,dae,dso,memo])
for sym in SYMS:
    if sym.get("skipToFall"):
        add_tree(sym, memo="이상없음 → 바로 의식/낙상")
    elif sym.get("direct"):
        add_tree(sym, dae=sym["direct"][0], dso=sym["direct"][1])
    elif sym.get("branch"):
        br=BRANCHES.get(sym["branch"],{})
        for opt in br.get("opts",[]):
            if opt.get("key"):
                add_tree(sym, q1=br.get("q",""), s1=opt["label"], dae=opt["key"][0], dso=opt["key"][1])
            elif opt.get("next"):
                br2=BRANCHES.get(opt["next"],{})
                for o2 in br2.get("opts",[]):
                    if o2.get("key"):
                        add_tree(sym, q1=br.get("q",""), s1=opt["label"],
                                 q2=br2.get("q",""), s2=o2["label"], dae=o2["key"][0], dso=o2["key"][1])
for c in t.iter_rows(min_row=2):
    for cell in c: cell.alignment=WRAP

# 시트2 사정세트
a = sheet("2.사정세트",
    ["대분류","소분류","과정","순서","진술문(마스터원문)","질문문구(💬)","입력형식","필수","비고"],
    [30,40,8,6,40,34,10,6,28])
PROC=[("사정 및 평가(A&E)","A&E"),("간호진단(D)","D"),("계획 및 중재(P&E)","P&E")]
for cdae, subs in DATA.items():
    for cdso, blocks in subs.items():
        merged={"A&E":[], "D":[], "P&E":[]}
        seen={"D":set(),"P&E":set()}
        for blk in blocks.values():
            for full,short in PROC:
                for it in blk.get(full,[]):
                    if short in ("D","P&E"):
                        if it["text"] in seen[short]: continue
                        seen[short].add(it["text"])
                    merged[short].append(it)
        for _,short in PROC:
            for i,it in enumerate(merged[short],1):
                txt=it["text"]
                q = QMAP.get(txt,"") if short=="A&E" else ""
                fmt = detect_fmt(txt) if short=="A&E" else ""
                a.append([cdae,cdso,short,i,txt,q,fmt,"",it.get("note","")])
for row in a.iter_rows(min_row=2):
    for cell in row: cell.alignment=WRAP

# 시트3 환자설명
e = sheet("3.환자설명", ["진단","말할 문구(📢)","연동 P&E"], [22,52,34])
for dx, items in EXPLAIN.items():
    for it in items:
        e.append([dx, it["say"], it["pe"]])
for row in e.iter_rows(min_row=2):
    for cell in row: cell.alignment=WRAP

# 시트4 도관
c = sheet("4.도관", ["도관 진술문"], [55])
for cat in CATS: c.append([cat])

# 시트5 낙상교육
f = sheet("5.낙상교육", ["id","말할 문구(📢)","기록 진술문(여러개는 | 로 구분)"], [12,50,50])
for it in EDU:
    f.append([it["id"], it["say"], " | ".join(it["records"])])
for row in f.iter_rows(min_row=2):
    for cell in row: cell.alignment=WRAP

wb.save(OUT)
print("스타터 엑셀 생성:", OUT)
print(f"  주호소트리 {t.max_row-1}행 / 사정세트 {a.max_row-1}행 / 환자설명 {e.max_row-1}행 / 도관 {c.max_row-1} / 낙상교육 {f.max_row-1}")

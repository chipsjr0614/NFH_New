#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
소아과교수_컨펌표_260811.xlsx 생성

무엇인가
  투약 앱 소아 탭에서 **임상 판단이 없어 확정 못 한 항목**을 소아과 교수 자문용
  엑셀로 만든다. 교수님이 「검토」 열에 ✅/✏️/❌ 만 찍고 「원내 확정값」을 적으면
  그대로 앱에 반영한다.

★ 데이터 무결성
  - 「참고값」은 전부 외부 표준 자료에서 찾아온 값이며 **앱에 적용하지 않았다.**
  - 근거가 없는 칸은 비우거나 「근거 없음」으로 둔다. 그럴듯한 값으로 채우지 않는다.
  - 정본 문서: 소아과교수_질의서_260811.md (이 스크립트는 그것을 엑셀로 옮긴 것)

형식
  주사제_작용부작용_검수표_260618.xlsx 의 집 스타일을 따른다.
  (제목 병합 → 빨간 경고 → 빈 줄 → 남색 헤더 → freeze A5)

사용:  python3 build_소아과컨펌표_260811.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
OUT  = os.path.join(BASE, '소아과교수_컨펌표_260811.xlsx')

# ── 집 스타일 (주사제 검수표 260618 과 동일) ──────────────────────────
HDR_FILL = PatternFill('solid', fgColor='FF305496')
HDR_FONT = Font(bold=True, size=11, color='FFFFFFFF')
TITLE    = Font(bold=True, size=13)
WARN     = Font(bold=False, size=10, color='FFC00000')
GRP_FILL = PatternFill('solid', fgColor='FFDCE6F1')   # 묶음 구분 줄
ASK_FILL = PatternFill('solid', fgColor='FFFFF2CC')   # 교수님이 채울 칸
RISK     = Font(bold=True, size=11, color='FFC00000') # ★ 임상 위험 표시
WRAP     = Alignment(wrap_text=True, vertical='center')
WRAP_T   = Alignment(wrap_text=True, vertical='top')
CTR      = Alignment(wrap_text=True, vertical='center', horizontal='center')
THIN     = Side(style='thin', color='FFBFBFBF')
BOX      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── 근거 코드 ────────────────────────────────────────────────────────
EVID = [
    ('A', 'PALS_소아응급계산기_DB_설계서.xlsx',
     '저장소 내 자료 (refs/)',
     '근거가 약물_참고문헌.md 에 AHA PALS 2020 / 2023 focused update 등으로 문서화됨',
     ''),
    ('B', 'UNC Pediatric IV Push Quick ED Reference Table (2025)',
     '미국 노스캐롤라이나대 병원 소아 응급 IV push 투약표',
     '소아 응급 정맥 푸시 약물의 용량·최대치·희석을 한 장으로 정리한 기관 표준표',
     'https://www.med.unc.edu/pediatrics/cccp/wp-content/uploads/sites/1156/2025/06/Pediatric-IV-Push-Quick-ED-Reference-Table.pdf'),
    ('C', 'OHSU Pediatric Emergency Management Guide 5th ed.',
     '미국 오리건보건과학대 소아 응급 포켓카드',
     '소아 응급 전반(소생·수액·지속주입·중독)을 담은 기관 표준 포켓카드',
     'https://www.ohsu.edu/sites/default/files/2019-04/DCH-3042167-Emergency-Pocket-Card-v5.pdf'),
    ('D', 'AAP (미국소아과학회) Pediatrics',
     '학회 공식 간행물',
     'Naloxone 소아 용량 / Vitamin K and the Newborn Infant (2022)',
     'https://publications.aap.org/pediatrics/article/149/3/e2021056036/184866/Vitamin-K-and-the-Newborn-Infant'),
    ('H', '소아응급_체중구간표_병원컨펌_260810.xlsx',
     '병원 제공 (refs/병원컨펌_소아_260810/)',
     '체중 11구간 × 장비·소생·RSI. ★ 검산으로 수치 오류 3건 검출 → Q1~Q3',
     ''),
    ('S', '입고예정약품 - 응급실요청 20260608.xlsx',
     '원내 재고 정본 (refs/)',
     '134건. 옴니셀·비품·현재입고량 포함. 원내 규격·농도의 1차 정본',
     ''),
    ('검산', 'AI 산술 검산',
     '이 표 작성 과정',
     '컨펌표 각 칸을 열 대표체중으로 나눠 mg/kg 역산 → 열 간 일관성 대조. ★ 임상 확정값 아님',
     ''),
    ('—', '근거 없음',
     '',
     '표준 투약표에서 찾지 못한 항목. ★ 지어내지 않고 비워 두었다',
     ''),
]

# ── 본 표 ────────────────────────────────────────────────────────────
# (Q번호, 묶음, 항목, 현재 앱 값, 참고값(초안), 근거, 쟁점, 위험★)
ROWS = [
 # ── Q1~Q3 : 병원 컨펌표 수치 오류 ──
 ('Q1', '① 병원 컨펌표 수치 오류', 'Epinephrine IV · 6-7kg 열',
  '앱은 컨펌표 값을 쓰지 않고 0.01mg/kg 공식으로 계산 중',
  '0.065mg (0.65ml)', 'H + 검산',
  '표 기재 = 0.65mg. 같은 행 다른 10개 열은 전부 0.010mg/kg인데 이 열만 0.100mg/kg(10배). '
  '괄호 안 ml값(0.65ml)은 0.1mg/ml 기준 0.065mg에 해당 → mg 표기만 어긋남', True),

 ('Q2', '① 병원 컨펌표 수치 오류', 'Cardioversion · 19-23 / 24-29 / 30-36kg 열',
  '★ 앱에 심율동전환 자체가 없음 (제세동만 탑재)',
  '21J/42J · 27J/53J · 33J/66J', 'H + 검산',
  '표 기재 = 40J/80J · 53J/106J · 66J/132J 로 바로 윗줄 제세동 값과 완전히 동일. '
  '나머지 8개 열에서는 제세동의 정확히 1/2 (1J/kg → 2J/kg)', True),

 ('Q3', '① 병원 컨펌표 수치 오류', 'Epinephrine ET 행 라벨',
  '★ 앱에 ET(기관내) 투여 경로가 없음',
  '행 이름을 1:1,000 으로 정정', 'H + 검산',
  'mg÷ml 역산 시 11개 열 전부 1.00mg/ml = 1:1,000. 라벨은 1:10,000. '
  '라벨대로 1:10,000을 조제해 표의 ml을 투여하면 실제 용량의 1/10', True),

 ('Q1-추가①', '① 병원 컨펌표 수치 오류', 'Ketamine · 3 / 4 / 5kg 열',
  '해당 없음', '', 'H',
  '표에 빈칸 — 해당 체중에서 사용하지 않는다는 뜻입니까, 누락입니까?', False),
 ('Q1-추가②', '① 병원 컨펌표 수치 오류', 'Intraosseous (IO) · 3 / 4 / 5kg 열',
  '앱에 「확인 필요」로 비워 둠', '', 'H',
  '표에 「?」 — 해당 체중 IO 바늘 규격', False),
 ('Q1-추가③', '① 병원 컨펌표 수치 오류', 'Succinylcholine · 15-18kg 열',
  '원내 미보유로 앱에서 제외', '패턴상 33mg', 'H + 검산',
  '표 기재 30mg — 의도된 값입니까? (원내 미보유라 앱 반영 계획은 없음)', False),

 # ── Q4~Q7 : 자료 간 값 불일치 ──
 ('Q4', '② 자료 간 값이 달라 못 고름', 'Methylprednisolone',
  '「원내 값 지정」으로 비워 둠',
  'A: 1~2mg/kg (max 125mg)\nB: 1~2mg/kg/day (max 60mg/day)\nC: 2mg/kg ×1 → 0.5~1mg/kg q6h (max 60mg/day, <12세)',
  'A·B·C',
  '초회 용량과 최대치가 모두 다름. 원내 Predisol 125mg vial (Hydrocortisone 미보유 → 대체 사용)', False),

 ('Q5', '② 자료 간 값이 달라 못 고름', 'Famotidine',
  '「원내 값 지정」으로 비워 둠',
  'A: 0.5~1mg/kg (max 50mg)\nB: 0.25mg/kg (max 40mg/day)', 'A·B',
  '2~4배 차이. 원내 Gaster 20mg vial', False),

 ('Q6', '② 자료 간 값이 달라 못 고름', 'Diazepam (경련)',
  '「원내 값 지정」으로 비워 둠',
  'A: PR 0.5mg/kg (max 20mg)\nB: IV 0.1~0.3mg/kg (max 10mg)\nC: IV 0.1~0.2mg/kg (max 10mg)',
  'A·B·C',
  '투여경로(PR ↔ IV)와 용량이 모두 다름. 원내는 IV 제형(10mg/2mL)만 보유', False),

 ('Q7', '② 자료 간 값이 달라 못 고름', 'Ipratropium 네뷸라이저',
  '「확인 필요」로 비워 둠',
  'C: Duoneb(Albuterol+Ipratropium 조합제) 3mL q20분 ×3', 'C',
  '★ 참고 자료는 조합 제형 기준인데 원내는 Atrovent UDV 500mcg/2mL 단독 제형 — '
  '조합제 용량을 그대로 쓸 수 없음. 단독 제형 소아 용량 지정 필요', False),

 # ── Q8~Q12 : 근거 미확보 ──
 ('Q8', '③ 표준표에 없어 근거 미확보', 'Budesonide (Pulmicort respule)',
  '「확인 필요」로 비워 둠', '', '—',
  'B·C 표에 없음. 크룹·천식 RCT에 2mg 사용례가 있으나 표준 투약표 근거가 아니라 채우지 않았음. '
  '원내 Pulmicort respule 0.5mg/2mL 보유 — 원내 크룹에 사용합니까?', False),

 ('Q9', '③ 표준표에 없어 근거 미확보', 'Acetylcysteine (Mucomyst)',
  '앱 해독 탭에 성인 150mg/kg loading 만 있음', '', '—',
  '소아 용량 미확보. 원내 Mucomyst 20% 4mL 보유. 소아 아세트아미노펜 중독 프로토콜 필요', False),

 ('Q10', '③ 표준표에 없어 근거 미확보', 'Potassium chloride',
  '「확인 필요」로 비워 둠', '', '—',
  '소아 응급 표준표에 없음(원내 희석 프로토콜 사안). 원내 KCl 15% 40mEq/20mL. ★ 반드시 희석 후 투여', False),

 ('Q11', '③ 표준표에 없어 근거 미확보', 'Vasopressin',
  '「확인 필요」로 비워 둠', '', 'B·C',
  'PALS가 소아 심정지 일상 권고로 두지 않음 → B·C 표에 없음. '
  '★ 앱에서 행을 삭제해도 되겠습니까? 원내 Ssopresin 20unit/1mL 비품 6', False),

 ('Q12', '③ 표준표에 없어 근거 미확보', 'Lacosamide (Vimsk)',
  '「확인 필요」로 비워 둠', '', '—',
  'B·C 표에 없음. 원내 Vimsk 200mg/20mL 응급실 비치 — 소아 경련에 사용합니까?', False),

 # ── Q13~Q19 : 현재 앱 값이 표준 자료와 어긋남 ──
 ('Q13', '④ 현재 앱 값이 어긋남', 'Ketorolac',
  '0.5mg/kg · max 30mg', 'B: 0.5mg/kg · max 15mg/회', 'A·B',
  'max 2배 차이 (A는 30mg)', False),

 ('Q14', '④ 현재 앱 값이 어긋남', 'Ondansetron',
  '0.15mg/kg · max 8mg', 'B: 0.1mg/kg (<40kg)', 'A·B',
  'A는 0.15mg/kg', False),

 ('Q15', '④ 현재 앱 값이 어긋남', 'Midazolam (진정)',
  '0.05~0.1mg/kg · ★ 상한 없음', 'B: max 2mg\nA: max 5mg', 'A·B',
  '★ 앱에 상한이 아예 없다 — 안전장치 부재. 원내 5mg/5mL · 15mg/3mL 두 규격', True),

 ('Q16', '④ 현재 앱 값이 어긋남', 'Atropine',
  '1회 max 0.5mg · ★ 총 최대치 없음',
  'B: 1회 max 0.5mg + 총 최대 소아 1mg / 청소년 3mg', 'B',
  '★ 앱에 총 최대치가 없다 — 반복 투여 시 상한 없음', True),

 ('Q17', '④ 현재 앱 값이 어긋남', 'Rocuronium',
  '1.2mg/kg', 'B: 1mg/kg (max 100mg)\nH: 1mg/kg', 'B·H',
  '병원 컨펌표와 UNC 모두 1mg/kg — 앱만 1.2mg/kg. 원내 응급실 유일 근이완제', False),

 ('Q18', '④ 현재 앱 값이 어긋남', 'Dexamethasone',
  '항구토 0.15mg/kg · max 10mg (크룹 적응증 없음)',
  '크룹: 0.6mg/kg · max 16mg', 'B·C',
  '★ 크룹 용량이 앱에 없다 — 추가할까요? 원내 5mg/1mL', False),

 ('Q19', '④ 현재 앱 값이 어긋남', '유지수액 공식',
  'Holliday-Segar', 'C: 4-2-1 rule', 'C',
  '✅ 동일 공식 — 확인만 부탁드립니다', False),

 # ── Q20~Q22 : 염 종류·희석 (약제부 공동) ──
 ('Q20', '⑤ 염·희석 임상 판단 (약제부 공동)', '칼슘',
  'Calcium gluconate 60mg/kg', '', 'H·S + 칩스 확인',
  '★ 컨펌표는 Calcium chloride 20mg/kg 기준인데 원내는 Ca gluconate 10% 2g/20mL 단독 보유(CaCl₂ 미비치). '
  '두 염은 원소칼슘이 약 3배 차이 (CaCl₂ 27% ↔ gluconate 9%). 원내 기준 mg/kg 지정 필요', True),

 ('Q21', '⑤ 염·희석 임상 판단 (약제부 공동)', 'Sodium bicarbonate',
  '1mEq/kg', '', 'H·S',
  '컨펌표는 4.2% 전제, 원내는 8.4% 20mEq/20mL. mEq는 같으나 투여 부피가 절반. '
  '소아에게 4.2%로 희석합니까, 8.4% 그대로입니까?', False),

 ('Q22', '⑤ 염·희석 임상 판단 (약제부 공동)', 'Atropine 희석',
  '0.02mg/kg (min 0.1mg · max 0.5mg)', '', 'H·S',
  '★ 컨펌표는 0.1mg/ml 전제, 원내는 0.5mg/1mL. 재고 앰플로 표의 ml을 그대로 뽑으면 5배 과량. '
  '희석액을 별도 조제합니까, 원액 기준으로 ml을 재계산합니까?', True),

 # ── Q23~Q26 : 수액 상한 (자료 자체가 없음) ──
 ('Q23', '⑥ 수액 상한 — 자료 미확보', '1회 볼루스 상한',
  '저혈량/아나필락시스 20ml/kg · 패혈증 10-20 · 심인성 5-10',
  '등장액 10~20ml/kg', 'A·C',
  '★ 병원 컨펌 파일에 수액 항목이 아예 없음. 원내 기준 ml/kg 지정 필요', False),

 ('Q24', '⑥ 수액 상한 — 자료 미확보', '누적 볼루스 상한',
  '쇼크 전략표에 「총 60ml/kg」 문구만 있고 누적 트래커 없음',
  '60ml/kg 후 강심제 고려', 'A·C',
  '상한값과 그 이후 경로 지정 필요 — 확정되면 앱에 누적 트래커 탑재', False),

 ('Q25', '⑥ 수액 상한 — 자료 미확보', '볼루스 감량 대상',
  '심인성 쇼크 5-10ml/kg 만 반영',
  '심근염 · DKA · 중증영양실조는 5~10ml/kg', 'C',
  '원내에 별도 감량 기준이 있습니까?', False),

 ('Q26', '⑥ 수액 상한 — 자료 미확보', 'DKA 초기 볼루스',
  '10-20ml/kg', 'C: 10ml/kg 단일', 'C',
  '어느 쪽입니까? ★ 인슐린·중탄산은 볼루스 금기 (C)', False),

 # ── Q27~Q29 : 소아 지속주입 ──
 ('Q27', '⑦ 소아 지속주입 (CIV)', 'Norepinephrine — 농도',
  '★ 계산 불가 (앱에 「확인 필요」)',
  '용량 범위 0.05~1 mcg/kg/min', 'A·C',
  '★ 농도 충돌 — 설계서는 소아 16mcg/mL, 성인 V4는 20mcg/mL로 확정. '
  '소아도 20mcg/mL로 통일합니까? 원내 Norpin 4mg/4mL', True),

 ('Q28', '⑦ 소아 지속주입 (CIV)', 'Dobutamine — 용량 범위',
  '★ 계산 불가 (농도는 확정, 범위 미정)',
  '농도 2,000mcg/mL (DoBUtamine/D5W 500mg/250mL 프리믹스)', 'S',
  'C에 소아 용량 범위 표기가 없음. 범위(mcg/kg/min)만 지정되면 즉시 계산 가능', False),

 ('Q29', '⑦ 소아 지속주입 (CIV)', 'Epinephrine — 희석법',
  '★ 계산 불가 (앱에 「확인 필요」)',
  '용량 범위 0.01~0.1 mcg/kg/min (쇼크)', 'C',
  '원내는 1mg/1mL 앰플만 보유 — 소아용 지속주입 희석 농도가 없어 ml/hr을 낼 수 없음', False),

 ('참고', '⑦ 소아 지속주입 (CIV)', 'Dopamine',
  '✅ 계산 중 — 2·5·10·15·20 mcg/kg/min 눈금표',
  '농도 2,000mcg/mL (Dopamix 400mg/200mL 프리믹스)\n범위 2~20 mcg/kg/min', 'S·C',
  '자료가 일치해 이미 반영했습니다 — 확인만 부탁드립니다', False),

 # ── 기타 ──
 ('기타①', '⑧ 그 밖에', '3% 고장성 식염수',
  '앱 미탑재', '두개내압 상승 시 3ml/kg (범위 2~5), 15~30분', 'C',
  '원내 보유·사용 여부 확인 대기', False),
 ('기타②', '⑧ 그 밖에', 'Racemic epinephrine',
  '앱 미탑재', '크룹 stridor 0.25mL ×2회 네뷸', 'C',
  '원내 미보유 → Ventolin + 스테로이드 대응이 맞습니까?', False),
 ('기타③', '⑧ 그 밖에', '기도 장비 — NPA 규격',
  '앱 3.5Fr (미확인 배지)', 'H: 14Fr', 'H',
  '★ 4배 차이로 판정 불가. 장비 실물 규격 목록이 필요합니다', True),
 ('기타④', '⑧ 그 밖에', '근이완제 운영',
  'Rocuronium 단독 (원내 미보유 5종 제외)', '', 'S + 칩스 결정 260810',
  'Succinylcholine 등 5행은 원내 미보유로 제외했습니다. 임상적으로 괜찮습니까?', False),
]

# ── 이미 반영한 초안 8건 (검수만) ────────────────────────────────────
DONE = [
 ('Naloxone', '0.1mg/kg IV (1회 max 2mg), 30초에 걸쳐, q2~3분 반복', 'B·C·D 일치',
  'D: 5세 미만 또는 20kg 이하는 0.1mg/kg / 20kg 초과는 2mg. 알칼리 용액과 혼합 금지. 원내 0.4mg/mL 실물 보유 확인(260810)'),
 ('Flumazenil', '0.01mg/kg (1회 max 0.2mg) IV q1분 · 누적 max 0.05mg/kg 또는 1mg 중 낮은 쪽', 'B·C 완전 일치',
  '★ 만성 벤조 복용자·경련 환자 금기. 의도적 과량복용에 일상적으로 쓰지 않음. 원내 Flunil 0.5mg/5mL'),
 ('활성탄', '0.5~1g/kg PO/NG (sorbitol 없는 제형)', 'C',
  '섭취 후 가급적 1시간 이내. 원내 New Heuk powder 50g 실물 보유 확인(260810)'),
 ('Mannitol', '0.25~1g/kg IV over 20분', 'C',
  '저혈량 아닌지 먼저 확인 · ★ 5μm 미만 필터 필요. 원내 20% 100mL'),
 ('Salbutamol 네뷸', '0.15mg/kg (최소 2.5mg, 통상 2.5~5mg) q20분 ×3 → 필요시 q1시간 · 지속 5~15mg/hr', 'A·C 일치',
  '원내 Ventolin Nebule 2.5mg/2.5mL'),
 ('Insulin (속효성)', 'DKA 0.05U/kg/hr 지속주입 (★볼루스 금지) · 고칼륨 0.1U/kg IV (max 10U) + Glucose 0.5g/kg 병용', 'C',
  'DKA에서 혈당 <300 → D5NS 추가, <200 → D10 추가. 원내 Humulin R 100IU/mL'),
 ('Dextrose 10%', 'Glucose 0.5g/kg IV (max 25g) → D10W 기준 5mL/kg', 'C',
  'B는 6개월 미만 0.25~0.5g/kg / 6개월 이상 0.5~1g/kg. 원내 D10W 1000mL bag'),
 ('Vitamin K1', '신생아 예방 0.5~1mg IM 1회 (출생체중 >1500g는 1mg)', 'D (AAP 2022)',
  '★ 체중 비례가 아닌 고정 용량 — 계산기에 kg 곱을 넣지 말 것. 원내 10mg/1mL'),
]


def style_header(ws, headers, widths, title, warn, freeze='A5'):
    ncol = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    ws.cell(1, 1, title).font = TITLE
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    c = ws.cell(2, 1, warn); c.font = WARN; c.alignment = WRAP
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 42
    for i, h in enumerate(headers, 1):
        c = ws.cell(4, i, h)
        c.font, c.fill, c.alignment, c.border = HDR_FONT, HDR_FILL, CTR, BOX
        ws.column_dimensions[get_column_letter(i)].width = widths[i-1]
    ws.row_dimensions[4].height = 34
    ws.freeze_panes = freeze


wb = Workbook()

# ══ 시트 1 — 검수 ════════════════════════════════════════════════════
ws = wb.active
ws.title = '소아과 검수'
HEAD = ['Q', '묶음', '항목', '현재 앱 값', '참고값 (초안 — 미적용)', '근거',
        '쟁점 — 왜 확정 못 했나', '검토\n✅채택 / ✏️수정 / ❌미사용', '원내 확정값', '메모']
style_header(ws, HEAD, [11, 22, 26, 30, 34, 12, 52, 15, 20, 22],
    '투약 앱 소아 계산기 — 소아과 교수 컨펌표  (2026-08-11)',
    '🛑 「참고값」은 전부 외부 표준 자료(AHA PALS 근거 설계서·UNC·OHSU·AAP)에서 찾은 초안이며 원내 컨펌이 아닙니다. '
    '앱에 적용하지 않고 비워 둔 상태입니다.  ▸ 「검토」·「원내 확정값」 두 열(노란 칸)만 채워 주시면 그대로 반영합니다.  '
    '▸ 근거 코드(A·B·C·D·H·S)의 원문 출처는 「근거」 시트 참조.  ▸ 재고·규격·희석 사실관계는 약제부 별도 질의 중입니다.')

r = 5
prev_grp = None
for q, grp, item, cur, ref, ev, issue, risk in ROWS:
    vals = [q, grp if grp != prev_grp else '', item, cur, ref, ev, issue, '', '', '']
    for i, v in enumerate(vals, 1):
        c = ws.cell(r, i, v)
        c.alignment = WRAP_T if i in (5, 7) else WRAP
        c.border = BOX
        if i in (8, 9):
            c.fill = ASK_FILL
        if grp != prev_grp and i == 2:
            c.fill = GRP_FILL; c.font = Font(bold=True, size=11)
    if risk:
        ws.cell(r, 3).font = RISK
    # 줄바꿈·본문 길이에 맞춰 행 높이
    lines = max(len(str(ref).split('\n')), 1)
    ws.row_dimensions[r].height = max(34, 15 * lines, min(96, 13 * (len(issue) // 40 + 1)))
    prev_grp = grp
    r += 1

dv = DataValidation(type='list', allow_blank=True,
                    formula1='"✅ 채택,✏️ 원내 값으로 수정,❌ 원내 미사용,보류"')
ws.add_data_validation(dv)
dv.add('H5:H%d' % (r - 1))
ws.auto_filter.ref = 'A4:J%d' % (r - 1)

# 맺음말
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
c = ws.cell(r, 1,
    '※ 회신 주시면 「원내 확정값」만 앱에 반영하고, 앱 화면의 🩺 Q배지를 제거합니다.  '
    '정본 문서: 소아과교수_질의서_260811.md · 소아약물_용량초안_검수표_260810.md')
c.font = Font(size=9, color='FF808080'); c.alignment = WRAP
ws.row_dimensions[r].height = 28

# ══ 시트 2 — 근거 ════════════════════════════════════════════════════
ws2 = wb.create_sheet('근거')
H2 = ['코드', '자료', '성격', '내용 / 비고', '링크']
style_header(ws2, H2, [8, 46, 30, 62, 60],
    '근거 자료 (Evidence)',
    '⚠️ AHA 원문(Circulation)은 유료 구독이라 직접 확인하지 못했습니다. '
    'A의 근거 표기와 B·C(미국 대학병원 기관 표준표)를 교차 대조하는 방식으로 갈음했습니다. '
    '따라서 이 표의 값은 「AHA 확정」이 아니라 「AHA 근거 자료 + 기관 표준표 교차 대조」입니다.')
r2 = 5
for code, name, kind, desc, url in EVID:
    for i, v in enumerate([code, name, kind, desc, url], 1):
        c = ws2.cell(r2, i, v); c.alignment = WRAP; c.border = BOX
    ws2.cell(r2, 1).font = Font(bold=True, size=11)
    if url:
        ws2.cell(r2, 5).hyperlink = url
        ws2.cell(r2, 5).font = Font(size=10, color='FF0563C1', underline='single')
    ws2.row_dimensions[r2].height = 42
    r2 += 1

# ══ 시트 3 — 이미 반영 (검수만) ══════════════════════════════════════
ws3 = wb.create_sheet('이미 반영 (검수만)')
H3 = ['#', '약물', '반영한 초안 용량', '근거', '주의 / 비고',
      '검토\n✅채택 / ✏️수정 / ❌미사용', '원내 확정값', '메모']
style_header(ws3, H3, [5, 20, 46, 16, 52, 15, 20, 22],
    '이미 앱에 반영한 초안 8건 — 검수만 부탁드립니다',
    '🛑 아래 8건은 자료가 일치해 앱에 「초안」 배지를 달고 넣어 둔 상태입니다. '
    '검수 후 배지를 뗍니다. 나머지 항목은 「소아과 검수」 시트에 있습니다.')
r3 = 5
for i, (name, dose, ev, note) in enumerate(DONE, 1):
    for j, v in enumerate([i, name, dose, ev, note, '', '', ''], 1):
        c = ws3.cell(r3, j, v); c.alignment = WRAP; c.border = BOX
        if j in (6, 7):
            c.fill = ASK_FILL
    ws3.row_dimensions[r3].height = 46
    r3 += 1
dv3 = DataValidation(type='list', allow_blank=True,
                     formula1='"✅ 채택,✏️ 원내 값으로 수정,❌ 원내 미사용,보류"')
ws3.add_data_validation(dv3)
dv3.add('F5:F%d' % (r3 - 1))

wb.save(OUT)

n_risk = sum(1 for x in ROWS if x[7])
print('✅ 생성: %s' % os.path.relpath(OUT, BASE))
print('   [소아과 검수] %d행  (★ 임상 위험 강조 %d건)' % (len(ROWS), n_risk))
print('   [근거]        %d건' % len(EVID))
print('   [이미 반영]   %d건' % len(DONE))

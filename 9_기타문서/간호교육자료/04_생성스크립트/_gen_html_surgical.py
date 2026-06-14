#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""외과계열 v1.1 xlsx → 응급실 외과계열 교육프로그램 v1.0 HTML (단일 오프라인 파일).
새 엔진 포맷(A~D, 이미지 열 없음) 파싱. 그림 생략."""
import openpyxl, re, json
from _paths import find, out_html

SRC = find("응급실_외과계열_교육케이스_v1.1.xlsx")
DST = out_html("응급실_외과계열_교육프로그램_v1.0.html")

# 카테고리 정렬 순서(접두사 기준): 외과 → 화상 → 정형 → 흉부 → 신경외과
PREFIX_ORDER = {'G':0, 'B':1, 'O':2, 'T':3, 'NS':4}

def parse_sheet(ws):
    def cell(r,c): return ws.cell(r,c).value
    title = (cell(1,1) or "")
    m = re.match(r'CASE\s*#?(\S+)\s*\|\s*(.+)', title.replace('  ',' ').strip())
    if not m: return None
    no, name = m.group(1), m.group(2).strip()
    parts = [p.strip() for p in str(cell(2,1) or "").split('|')]
    cat  = parts[0] if parts else ""
    diff = parts[1].replace('난이도:','').strip() if len(parts)>1 else ""
    kw   = parts[2] if len(parts)>2 else ""
    mraw = str(cell(3,1) or "")
    mentor = mraw.split('\n',1)[1].strip() if '\n' in mraw else ""
    sections=[]; redflags=[]; memo=""; cur=None; mode='sec'
    for r in range(5, ws.max_row+1):
        A=cell(r,1); B=cell(r,2); C=cell(r,3); D=cell(r,4)
        if A:
            As=str(A).strip()
            if As.startswith('🔴') or 'Red Flag' in As: mode='red'; continue
            if As.startswith('⚡'):
                memo = As.split('|',1)[1].strip() if '|' in As else As.replace('⚡','').replace('핵심 암기','').strip()
                mode='memo'; continue
            if mode=='red' and As.startswith('⚠'): redflags.append(As.lstrip('⚠️ ').strip()); continue
            # 섹션 머리행: A만 있고 B 비어있음(병합셀) → 이모지 종류와 무관하게 감지
            if (B is None or str(B).strip()=='') and mode not in ('red','memo'):
                cur={"title":As,"items":[]}; sections.append(cur); mode='sec'; continue
        if mode=='sec' and B and cur is not None:
            cur["items"].append({"item":str(B).strip(),"content":str(C).strip() if C else "",
                                 "basis":str(D).strip() if D else "","img":""})
    if not sections: return None
    return {"no":no,"name":name,"cat":cat,"diff":diff,"kw":kw,"mentor":mentor,
            "sections":sections,"redflags":redflags,"memo":memo}

def sortkey(c):
    m = re.match(r'([A-Za-z]+)(\d+)', c["no"])
    pre, num = (m.group(1), int(m.group(2))) if m else ('Z', 999)
    return (PREFIX_ORDER.get(pre, 90), num)

wb = openpyxl.load_workbook(SRC, data_only=True)
seen=set(); cases=[]
for sn in wb.sheetnames:
    if sn.startswith('📋'): continue
    c = parse_sheet(wb[sn])
    if not c: continue
    if c["no"] in seen: continue
    seen.add(c["no"]); cases.append(c)
cases.sort(key=sortkey)
print(f"케이스 {len(cases)}개")

# 외과계열 SVG 도해를 섹션에 주입
from _svg_lib import SVG, SVG_MAP_SURG
_nimg=0
for c in cases:
    mp = SVG_MAP_SURG.get(c["no"], {})
    for s in c["sections"]:
        sid=None
        for key,vid in mp.items():
            if key in s["title"]: sid=vid; break
        s["svg"] = SVG.get(sid,"") if sid else ""
        if s["svg"]: _nimg+=1
print(f"삽입된 그림 {_nimg}개")

DATA = json.dumps(cases, ensure_ascii=False)

HTML = r'''<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>응급실 외과계열 교육 프로그램 v1.0 — 국립소방병원</title>
<style>
:root{
  --bg:#0d1117;--surf:#161b22;--surf2:#1c2330;--brd:#2a3441;
  --txt:#e6edf3;--mute:#8b98a8;--dim:#586273;
  --gold:#e3b341;
  --why:#4d9fff;--whybg:#0d2440;--whybd:#1e4976;
  --red:#ef4444;--redbg:#2a1012;--redbd:#7f1d1d;
  --c0:#1565C0;--c1:#2e7d32;--c2:#e65100;--c3:#8e24aa;--c4:#c62828;--c5:#00838f;
}
body.light{
  --bg:#eef2f6;--surf:#ffffff;--surf2:#f5f8fb;--brd:#dde5ee;
  --txt:#16202c;--mute:#5b6876;--dim:#9aa7b5;
  --why:#1565C0;--whybg:#eaf2fd;--whybd:#bcd6f5;
  --redbg:#fdeced;--redbd:#f5b5b5;
}
*{margin:0;padding:0;box-sizing:border-box;-webkit-tap-highlight-color:transparent;}
body{background:var(--bg);color:var(--txt);font-size:18px;line-height:1.7;
  font-family:-apple-system,BlinkMacSystemFont,'Apple SD Gothic Neo','Noto Sans KR',sans-serif;
  transition:background .2s,color .2s;}
.wrap{max-width:880px;margin:0 auto;padding:0 18px;}

/* 상단바 */
.top{position:sticky;top:0;z-index:40;background:var(--surf);border-bottom:1px solid var(--brd);}
.top-in{max-width:880px;margin:0 auto;padding:12px 18px;display:flex;align-items:center;gap:12px;}
.brand{font-size:18px;font-weight:800;letter-spacing:-.4px;cursor:pointer;white-space:nowrap;}
.brand em{color:var(--red);font-style:normal;}
.top .sp{flex:1;}
.tgl{display:flex;gap:5px;}
.tgl button{padding:7px 12px;border:1px solid var(--brd);background:var(--surf2);color:var(--mute);
  border-radius:18px;font-size:13px;font-weight:600;cursor:pointer;font-family:inherit;}
.tgl button.on{background:var(--red);border-color:var(--red);color:#fff;}

/* 홈 */
#home{padding:20px 0 60px;}
.hero-home{text-align:center;padding:14px 0 18px;}
.hero-home h1{font-size:27px;font-weight:800;margin-bottom:6px;}
.hero-home p{color:var(--mute);font-size:15px;}
.prog{margin:14px auto 0;max-width:420px;background:var(--surf);border:1px solid var(--brd);border-radius:12px;padding:12px 16px;}
.prog .pbar{height:8px;background:var(--surf2);border-radius:5px;overflow:hidden;margin-top:8px;}
.prog .pfill{height:100%;background:linear-gradient(90deg,#2e7d32,#43a047);transition:width .4s;}
.search{position:relative;margin:18px 0 8px;}
.search input{width:100%;background:var(--surf);border:1.5px solid var(--brd);border-radius:12px;
  padding:14px 16px 14px 46px;color:var(--txt);font-size:16px;font-family:inherit;outline:none;}
.search input:focus{border-color:var(--why);}
.search .si{position:absolute;left:15px;top:50%;transform:translateY(-50%);font-size:20px;}

.cat{margin-top:26px;}
.cat-h{font-size:17px;font-weight:800;margin-bottom:12px;display:flex;align-items:center;gap:8px;}
.cat-h .cnt{font-size:13px;color:var(--mute);font-weight:600;}
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(250px,1fr));gap:11px;}
.card{background:var(--surf);border:1px solid var(--brd);border-radius:13px;padding:15px;cursor:pointer;
  transition:.15s;position:relative;border-left:4px solid var(--cc,#666);}
.card:hover{transform:translateY(-2px);border-color:var(--why);}
.card .no{font-size:12px;color:var(--mute);font-weight:700;}
.card .nm{font-size:16.5px;font-weight:800;margin:3px 0 7px;line-height:1.35;}
.card .df{font-size:13px;color:var(--gold);}
.card .vchk{position:absolute;top:13px;right:13px;font-size:16px;display:none;}
.card.viewed .vchk{display:block;}
.card.viewed{opacity:.78;}
.noresult{text-align:center;color:var(--mute);padding:40px 0;}

/* 케이스 상세 */
#detail{display:none;padding-bottom:70px;}
.back{display:inline-flex;align-items:center;gap:6px;margin:16px 0 4px;background:var(--surf);
  border:1px solid var(--brd);color:var(--txt);padding:9px 15px;border-radius:20px;font-size:14px;
  font-weight:600;cursor:pointer;font-family:inherit;}
.hero{color:#fff;border-radius:18px;padding:24px 22px;margin:8px 0 18px;}
.hero .cno{font-size:14px;font-weight:700;opacity:.9;letter-spacing:.5px;}
.hero h1{font-size:30px;font-weight:800;margin:4px 0 12px;letter-spacing:-.5px;line-height:1.2;}
.hero .meta{display:flex;flex-wrap:wrap;gap:8px;}
.chip{background:rgba(255,255,255,.16);border:1px solid rgba(255,255,255,.28);
  padding:5px 12px;border-radius:16px;font-size:13.5px;font-weight:600;}

.mentor{background:var(--surf2);border:1px solid var(--brd);border-left:4px solid var(--gold);
  border-radius:12px;padding:17px 19px;margin-bottom:20px;}
.mentor .mh{font-size:15px;font-weight:800;color:var(--gold);margin-bottom:7px;}
.mentor p{font-size:16.5px;}

.tabs{display:flex;gap:6px;overflow-x:auto;padding-bottom:10px;scrollbar-width:none;}
.tabs::-webkit-scrollbar{display:none;}
.tab{flex-shrink:0;padding:10px 15px;border-radius:20px;border:1px solid var(--brd);background:var(--surf);
  color:var(--mute);font-size:15px;font-weight:700;cursor:pointer;font-family:inherit;white-space:nowrap;transition:.15s;}
.tab.active{color:#fff;border-color:transparent;}

.panel{display:none;animation:fade .25s ease;}
.panel.active{display:block;}
@keyframes fade{from{opacity:0;transform:translateY(6px)}to{opacity:1;transform:none}}
.sec-head{display:flex;align-items:center;gap:9px;margin:14px 0 12px;}
.sec-head h2{font-size:20px;font-weight:800;}
.item{background:var(--surf);border:1px solid var(--brd);border-radius:13px;padding:18px;margin-bottom:13px;}
.item-title{font-size:17.5px;font-weight:800;margin-bottom:10px;display:flex;align-items:center;gap:8px;}
.item-title::before{content:"";width:8px;height:8px;border-radius:50%;background:var(--red);flex-shrink:0;}
.item-body{font-size:16px;white-space:pre-line;}
.why{margin-top:12px;background:var(--whybg);border:1px solid var(--whybd);border-radius:10px;padding:13px 15px;}
.why .wl{font-size:13px;font-weight:800;color:var(--why);letter-spacing:.5px;margin-bottom:5px;}
.why .wt{font-size:15px;white-space:pre-line;}
.fig{background:#fff;border:1px solid var(--brd);border-radius:14px;padding:14px;margin:6px 0 16px;box-shadow:0 1px 5px rgba(0,0,0,.07);}
.fig svg{width:100%;height:auto;display:block;}

.redbox{background:var(--redbg);border:1px solid var(--redbd);border-radius:14px;padding:18px;margin:22px 0 14px;}
.redbox h2{color:var(--red);font-size:19px;font-weight:800;margin-bottom:12px;}
.rf{display:flex;gap:10px;padding:11px 0;border-bottom:1px dashed var(--redbd);font-size:16px;}
.rf:last-child{border:none;}
.rf .ri{color:var(--red);flex-shrink:0;}

.memo{background:linear-gradient(135deg,#1A237E,#283593);color:#fff;border-radius:14px;padding:20px;margin-bottom:14px;}
.memo .ml{font-size:15px;font-weight:800;color:var(--gold);margin-bottom:8px;}
.memo p{font-size:16.5px;font-weight:600;}

.navbtns{display:flex;gap:10px;margin-top:18px;}
.navbtns button{flex:1;padding:13px;border:1px solid var(--brd);background:var(--surf);color:var(--txt);
  border-radius:12px;font-size:15px;font-weight:700;cursor:pointer;font-family:inherit;}
.navbtns button:disabled{opacity:.4;cursor:default;}
</style>
</head>
<body>

<div class="top"><div class="top-in">
  <div class="brand" onclick="goHome()"><em>ER</em> 외과계열 교육</div>
  <div class="sp"></div>
  <div class="tgl">
    <button id="bd" class="on" onclick="theme('dark')">🌙</button>
    <button id="bl" onclick="theme('light')">☀️</button>
  </div>
</div></div>

<div class="wrap">
  <!-- 홈 -->
  <div id="home">
    <div class="hero-home">
      <h1>🏥 응급실 외과계열 교육 프로그램</h1>
      <p>외과 · 화상 · 정형외과 · 흉부외과 · 신경외과 · 케이스 <span id="totalN"></span>개</p>
      <div class="prog">
        <div style="display:flex;justify-content:space-between;font-size:14px;font-weight:600">
          <span>학습 진행</span><span id="progTxt">0 / 0</span>
        </div>
        <div class="pbar"><div class="pfill" id="pfill" style="width:0%"></div></div>
      </div>
    </div>
    <div class="search">
      <span class="si">🔍</span>
      <input id="q" type="text" placeholder="병명·증상·키워드 검색…" oninput="renderHome()">
    </div>
    <div id="catList"></div>
  </div>

  <!-- 상세 -->
  <div id="detail"></div>
</div>

<script>
const CASES = ''' + DATA + r''';
const CAT_COLOR = {
  '🔪 외과/외상':'#37474F','🔥 화상':'#BF360C','🦴 정형외과':'#4E342E',
  '🫁 흉부외과':'#01579B','🧠 신경외과':'#283593'};
const SEC_COLOR = ['#1565C0','#2e7d32','#e65100','#8e24aa','#c62828','#00838f'];
const esc = s => (s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
let VIEWED = new Set(JSON.parse(localStorage.getItem('erSurgViewed')||'[]'));

function theme(t){
  document.body.classList.toggle('light',t==='light');
  document.getElementById('bd').classList.toggle('on',t==='dark');
  document.getElementById('bl').classList.toggle('on',t==='light');
  localStorage.setItem('erSurgTheme',t);
}
if(localStorage.getItem('erSurgTheme')==='light') theme('light');

function catColor(cat){ return CAT_COLOR[cat] || '#666'; }
function updateProg(){
  const n=VIEWED.size, t=CASES.length;
  document.getElementById('progTxt').textContent = n+' / '+t;
  document.getElementById('pfill').style.width = (t? n/t*100:0)+'%';
}
function renderHome(){
  const q=(document.getElementById('q').value||'').toLowerCase().trim();
  const list=CASES.filter(c=> !q || (c.name+c.kw+c.cat).toLowerCase().includes(q));
  const cats={};
  list.forEach(c=>{(cats[c.cat]=cats[c.cat]||[]).push(c);});
  const el=document.getElementById('catList');
  if(!list.length){ el.innerHTML='<div class="noresult">검색 결과가 없어요 🔍</div>'; return; }
  let html='';
  Object.keys(cats).forEach(cat=>{
    html+=`<div class="cat"><div class="cat-h">${esc(cat)} <span class="cnt">${cats[cat].length}개</span></div><div class="grid">`;
    cats[cat].forEach(c=>{
      html+=`<div class="card ${VIEWED.has(c.no)?'viewed':''}" style="--cc:${catColor(c.cat)}" onclick="openCase('${c.no}')">
        <span class="vchk">✅</span>
        <div class="no">CASE ${esc(c.no)}</div>
        <div class="nm">${esc(c.name)}</div>
        <div class="df">${esc(c.diff)}</div>
      </div>`;
    });
    html+='</div></div>';
  });
  el.innerHTML=html;
}
function openCase(no){
  const i=CASES.findIndex(c=>c.no===no); if(i<0) return;
  const c=CASES[i];
  VIEWED.add(no); localStorage.setItem('erSurgViewed',JSON.stringify([...VIEWED])); updateProg();
  const col=catColor(c.cat);
  let tabs='', panels='';
  c.sections.forEach((s,si)=>{
    const sc=SEC_COLOR[si%SEC_COLOR.length];
    tabs+=`<button class="tab ${si===0?'active':''}" style="${si===0?'background:'+sc:''}" onclick="selTab(this,'pn${si}','${sc}')">${esc(s.title)}</button>`;
    let items='';
    s.items.forEach(it=>{
      items+=`<div class="item"><div class="item-title">${esc(it.item)}</div>
        <div class="item-body">${esc(it.content)}</div>
        ${it.basis?`<div class="why"><div class="wl">💡 왜?</div><div class="wt">${esc(it.basis)}</div></div>`:''}</div>`;
    });
    const fig = s.svg?`<div class="fig">${s.svg}</div>`:'';
    panels+=`<div class="panel ${si===0?'active':''}" id="pn${si}">
      <div class="sec-head"><h2>${esc(s.title)}</h2></div>${fig}${items}</div>`;
  });
  const rf = c.redflags.length? `<div class="redbox"><h2>🔴 Red Flag — 절대 놓치면 안 되는 것</h2>
    ${c.redflags.map(r=>`<div class="rf"><span class="ri">⚠️</span><span>${esc(r)}</span></div>`).join('')}</div>`:'';
  const memo = c.memo? `<div class="memo"><div class="ml">⚡ 핵심 암기</div><p>${esc(c.memo)}</p></div>`:'';
  const prev = i>0? CASES[i-1]:null, next = i<CASES.length-1? CASES[i+1]:null;
  document.getElementById('detail').innerHTML = `
    <button class="back" onclick="goHome()">← 목록으로</button>
    <div class="hero" style="background:linear-gradient(135deg,${col}cc,${col})">
      <div class="cno">CASE ${esc(c.no)} · ${esc(c.cat)}</div>
      <h1>${esc(c.name)}</h1>
      <div class="meta"><span class="chip">난이도 ${esc(c.diff)}</span>${
        c.kw.split('/').map(k=>k.trim()).filter(k=>k).map(k=>`<span class="chip">${esc(k)}</span>`).join('')}</div>
    </div>
    ${c.mentor?`<div class="mentor"><div class="mh">💬 선배 말씀 (프셉마음)</div><p>${esc(c.mentor)}</p></div>`:''}
    <div class="tabs">${tabs}</div>
    ${panels}
    ${rf}${memo}
    <div class="navbtns">
      <button ${prev?'':'disabled'} onclick="${prev?`openCase('${prev.no}')`:''}">← ${prev?esc(prev.name):'이전'}</button>
      <button ${next?'':'disabled'} onclick="${next?`openCase('${next.no}')`:''}">${next?esc(next.name):'다음'} →</button>
    </div>`;
  document.getElementById('home').style.display='none';
  document.getElementById('detail').style.display='block';
  window.scrollTo(0,0);
}
function selTab(btn,pid,color){
  document.querySelectorAll('.tab').forEach(t=>{t.classList.remove('active');t.style.background='';});
  document.querySelectorAll('.panel').forEach(p=>p.classList.remove('active'));
  btn.classList.add('active'); btn.style.background=color;
  document.getElementById(pid).classList.add('active');
}
function goHome(){
  document.getElementById('detail').style.display='none';
  document.getElementById('home').style.display='block';
  renderHome(); window.scrollTo(0,0);
}
document.getElementById('totalN').textContent=CASES.length;
updateProg(); renderHome();
</script>
</body>
</html>'''

with open(DST,'w',encoding='utf-8') as f:
    f.write(HTML)
print(f"저장: {DST} ({len(HTML):,} bytes, ~{len(HTML)//1024}KB)")

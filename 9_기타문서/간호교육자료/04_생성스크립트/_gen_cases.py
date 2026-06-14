#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
응급실 내과 교육케이스 — 케이스 시트 자동 생성기
원본(v1.0) 복사 → 새 케이스 시트 추가 → v1.1로 저장
원본 절대 수정 금지 (read만), 결과는 새 파일명.
"""
import math, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from _paths import find, out_final
SRC = find("응급실_내과_교육케이스_v1.0.xlsx")
DST = out_final("응급실_내과_교육케이스_v1.1.xlsx")

# ── 색상 팔레트 ──────────────────────────────────────────────
SEC_COLORS = ["1565C0", "1B5E20", "E65100", "6A1B9A", "B71C1C", "00838F"]  # 순환
REDFLAG_HDR = "B71C1C"; REDFLAG_BODY = "FFEBEE"
MEMO_FILL = "1A237E"; HEADER_FILL = "1A2A3A"; MENTOR_FILL = "FFFDE7"
CAT_COLOR = {
    "❤️ 심장/순환기": "B71C1C", "🫁 호흡기": "00838F", "🧠 신경계": "1A237E",
    "🫀 소화기": "5D4037", "🩺 신장/비뇨기": "00695C", "🍬 내분비/대사": "E65100",
    "🦠 감염/패혈증": "1B5E20", "😰 기타 내과": "455A64",
}

def F(rgb): return PatternFill("solid", fgColor=rgb)
def white(sz, b=True): return Font(name="맑은 고딕", size=sz, bold=b, color="FFFFFF")
def dark(sz, b=False, color="212121"): return Font(name="맑은 고딕", size=sz, bold=b, color=color)
AL_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
AL_LT = Alignment(horizontal="left", vertical="top", wrap_text=True)
AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)

def est_lines(text, width):
    if not text: return 1
    n = 0
    for ln in str(text).split("\n"):
        n += max(1, math.ceil(len(ln) / width))
    return n

def render_case(wb, case):
    no, name, cat = case["no"], case["name"], case["cat"]
    diff, kw, mentor = case["diff"], case["kw"], case["mentor"]
    sheet_name = (f"{no}_{name}")[:31].replace("/", "·").replace("\\", "·").replace("?", "").replace("*", "").replace("[", "(").replace("]", ")")
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    catfill = CAT_COLOR.get(cat, "455A64")

    # 열너비
    for col, w in zip("ABCDEFG", [5, 22, 38, 38, 26, 2, 36]):
        ws.column_dimensions[col].width = w

    # R1 제목
    ws.merge_cells("A1:E1")
    ws["A1"] = f"CASE #{no}  |  {name}"
    ws["A1"].fill = F(catfill); ws["A1"].font = white(14, True); ws["A1"].alignment = AL_L
    ws.row_dimensions[1].height = 36
    # R2 메타
    ws.merge_cells("A2:E2")
    ws["A2"] = f"{cat}  |  난이도: {diff}  |  {kw}"
    ws["A2"].fill = F(catfill); ws["A2"].font = white(9, False); ws["A2"].alignment = AL_L
    ws.row_dimensions[2].height = 18
    # R3 선배 말씀
    ws.merge_cells("A3:E3")
    ws["A3"] = f"💬 선배 말씀 (프셉마음)\n{mentor}"
    ws["A3"].fill = F(MENTOR_FILL); ws["A3"].font = dark(9.5); ws["A3"].alignment = AL_LT
    ws.row_dimensions[3].height = 55
    # R4 표머리
    for col, txt in zip("ABCDE", ["섹션", "항목", "내용", "이론적 근거", "이미지"]):
        c = ws[f"{col}4"]; c.value = txt; c.fill = F(HEADER_FILL); c.font = white(10, True); c.alignment = AL_C
    ws.row_dimensions[4].height = 22

    r = 5
    for si, sec in enumerate(case["sections"]):
        color = SEC_COLORS[si % len(SEC_COLORS)]
        ws.merge_cells(f"A{r}:E{r}")
        ws[f"A{r}"] = f"  {sec['title']}"
        ws[f"A{r}"].fill = F(color); ws[f"A{r}"].font = white(10, True); ws[f"A{r}"].alignment = AL_L
        ws.row_dimensions[r].height = 22
        r += 1
        body_start = r
        rows = sec["rows"]
        for ri, row in enumerate(rows):
            item, content, basis = row.get("item", ""), row.get("content", ""), row.get("basis", "")
            ws[f"B{r}"] = item; ws[f"B{r}"].font = dark(9.5, True); ws[f"B{r}"].alignment = AL_LT
            ws[f"C{r}"] = content; ws[f"C{r}"].font = dark(9.5); ws[f"C{r}"].alignment = AL_LT
            ws[f"D{r}"] = basis; ws[f"D{r}"].font = dark(9.0); ws[f"D{r}"].alignment = AL_LT
            lines = max(est_lines(content, 36), est_lines(basis, 40), est_lines(item, 11))
            ws.row_dimensions[r].height = min(max(lines * 15 + 6, 30), 220)
            r += 1
            if ri < len(rows) - 1:
                ws.row_dimensions[r].height = 8
                r += 1
        # 이미지 칸 병합 + 캡션
        if sec.get("img"):
            if r - 1 > body_start:
                ws.merge_cells(f"E{body_start}:E{r-1}")
            ws[f"E{body_start}"] = sec["img"]
            ws[f"E{body_start}"].font = dark(8.5, False, "607D8B"); ws[f"E{body_start}"].alignment = AL_LT
        # 섹션 간 간격
        ws.row_dimensions[r].height = 8; r += 1

    # Red Flag
    ws.merge_cells(f"A{r}:E{r}")
    ws[f"A{r}"] = "  🔴 Red Flag — 절대 놓치면 안 되는 것"
    ws[f"A{r}"].fill = F(REDFLAG_HDR); ws[f"A{r}"].font = white(10, True); ws[f"A{r}"].alignment = AL_L
    ws.row_dimensions[r].height = 22; r += 1
    for rf in case["redflags"]:
        ws.merge_cells(f"A{r}:E{r}")
        ws[f"A{r}"] = f"  ⚠️  {rf}"
        ws[f"A{r}"].fill = F(REDFLAG_BODY); ws[f"A{r}"].font = dark(9.5, True, REDFLAG_HDR); ws[f"A{r}"].alignment = AL_L
        ws.row_dimensions[r].height = max(est_lines(rf, 90) * 15 + 6, 20); r += 1
    ws.row_dimensions[r].height = 8; r += 1

    # 핵심 암기
    ws.merge_cells(f"A{r}:E{r}")
    ws[f"A{r}"] = f"  ⚡ 핵심 암기  |  {case['memo']}"
    ws[f"A{r}"].fill = F(MEMO_FILL); ws[f"A{r}"].font = white(9.5, True); ws[f"A{r}"].alignment = AL_L
    ws.row_dimensions[r].height = max(est_lines(case["memo"], 95) * 15 + 10, 28)

    return sheet_name


def update_list(wb, completed):
    """케이스목록 시트에서 완성 케이스 상태/시트명/배경 갱신"""
    wl = wb["📋 케이스 목록"]
    # 헤더 R4, 데이터 R5부터. No NN → row 4+int(NN)
    for no, sheet_name in completed.items():
        row = 4 + int(no)
        wl.cell(row, 6).value = "✅ 완성"
        wl.cell(row, 7).value = sheet_name
        for col in range(1, 8):
            wl.cell(row, col).fill = F("FFCDD2")


# ════════════════════════════════════════════════════════════════
# 케이스 콘텐츠 — ❤️ 심장/순환기 (남은 9개)
# ════════════════════════════════════════════════════════════════
CAT = "❤️ 심장/순환기"
CASES = []

CASES.append({
 "no":"02","name":"급성 심근경색 (NSTEMI)","cat":CAT,"diff":"★★★☆☆",
 "kw":"부분폐색 / ST하강·T역전 / Troponin↑ / GRACE / 위험도층화 PCI",
 "mentor":'"STEMI는 \'당장 뚫어\'지만 NSTEMI는 \'위험한 만큼 빨리\'야. ST가 안 올라갔다고 안심하면 안 돼. Troponin 오르고 ST 내려가 있으면 이미 심근은 죽고 있는 거야. GRACE 점수로 얼마나 급한지 판단해."',
 "sections":[
   {"title":"📖 정의","img":"🖼️ 부분 폐색 관상동맥 단면 — 혈전이 내강 일부만 막은 모습","rows":[
     {"item":"한 줄 정의","content":"관상동맥이 \'부분\' 폐색되어 심근 일부가 허혈·괴사되는 상태. ST 상승은 없지만 Troponin이 오른다.","basis":"혈류가 완전히 끊긴 게 아니라 줄어든 상태 → 심내막하(subendocardial) 괴사 위주 → ST 상승 대신 하강·T 역전."},
     {"item":"STEMI와 차이","content":"STEMI: 완전폐색·ST상승·즉시 PCI(90분)\nNSTEMI: 부분폐색·ST하강/정상·위험도 따라 PCI(보통 24~72h)","basis":"둘 다 Troponin↑(심근 괴사 동반). 차이는 폐색 정도와 긴급도. NSTEMI도 고위험군은 24h 내 조기 침습."},
     {"item":"UA와 차이","content":"불안정협심증(UA): 같은 증상이지만 Troponin 정상(괴사 없음)\nNSTEMI: Troponin 상승(괴사 있음)","basis":"UA와 NSTEMI는 ECG·증상이 같음. Troponin 상승 여부가 갈림길 → Serial Troponin 필수."},
   ]},
   {"title":"🔬 병태생리","img":"🖼️ 불안정 플라크 파열 → 혈소판 혈전 형성 단계도","rows":[
     {"item":"진행 과정","content":"불안정 플라크 파열 → 혈전 형성 → 내강 부분 폐색 → 심내막하 허혈","basis":"심근벽은 안쪽(내막)이 혈류 마지막에 도달 → 부분 폐색 시 내막부터 손상 → ST 하강으로 표현."},
   ]},
   {"title":"😣 증상","img":"","rows":[
     {"item":"증상","content":"흉통·압박감(STEMI보다 약하거나 단속적일 수 있음), 방사통, 식은땀, 호흡곤란, 오심","basis":"허혈 심근의 통증 기전은 STEMI와 동일. 단 부분 폐색이라 강도·지속이 변동적일 수 있어 놓치기 쉬움."},
     {"item":"비전형","content":"당뇨·고령·여성은 흉통 없이 \'기운 없음·소화불량\'만 호소","basis":"자율신경 둔화 → 통증 인지 저하. 비전형 증상도 ECG·Troponin 먼저 확인."},
   ]},
   {"title":"📈 검사","img":"🖼️ NSTEMI 12-lead — ST 하강·T파 역전 강조","rows":[
     {"item":"ECG","content":"ST 분절 하강(≥0.5mm) 또는 T파 역전. 정상일 수도 있음.","basis":"심내막하 허혈 → 재분극 이상이 ST 상승이 아닌 \'하강\'으로 나타남. 정상 ECG도 NSTEMI 배제 못 함."},
     {"item":"Serial Troponin","content":"0h·3~6h 재검. 상승·변화 패턴 확인.","basis":"초기 음성 가능 → 시간 경과 후 상승(rising pattern)이 진단. 한 번 정상으로 배제 금지."},
     {"item":"GRACE 위험도","content":"나이·HR·SBP·Cr·Killip·심정지·ECG·Troponin으로 점수화 → 침습 시기 결정","basis":"고위험(GRACE>140)은 24h 내 조기 침습이 예후 개선. 위험도 층화가 NSTEMI 관리의 핵심."},
   ]},
   {"title":"💊 치료 근거","img":"","rows":[
     {"item":"즉시 처치","content":"12-lead ECG, IV 2개, 채혈(Troponin·CBC·BMP·Coag), 지속 모니터링, 산소(SpO₂<90%)","basis":"STEMI 가능성 배제·위험도 평가 동시 진행. 산소 과투여는 심근 손상 악화로 제한적."},
     {"item":"항혈소판","content":"Aspirin 300mg 씹어서 + P2Y12 억제제(Clopidogrel/Ticagrelor)","basis":"DAPT로 혈전 진행 차단. 두 경로(COX-1, P2Y12) 동시 억제."},
     {"item":"항응고","content":"Heparin / Enoxaparin / Fondaparinux","basis":"혈전 추가 형성 억제. PCI 전후 출혈위험 고려해 선택."},
     {"item":"항허혈","content":"NTG(SBP≥90), β차단제, 통증 시 Morphine","basis":"전부하·심박수↓ → 심근 산소요구량↓. NSTEMI는 즉시 PCI가 아니라 약물로 안정화 후 위험도 따라 PCI."},
   ]},
 ],
 "redflags":[
   "Troponin 상승 + ST 하강 → 이미 괴사 진행 중, STEMI 못지않게 위험",
   "흉통 지속 + 혈역학 불안정 → 즉시 침습(긴급 PCI) 대상",
   "초기 Troponin 음성이라고 귀가 금지 → Serial 재검 필수",
   "GRACE 고위험 / 재발성 흉통 / 새 ST변화 → 24h 내 조기 침습 보고",
 ],
 "memo":"부분폐색·ST하강/T역전 / Troponin Serial(0·3~6h) / DAPT+항응고 / GRACE로 PCI 시기 결정",
})

CASES.append({
 "no":"03","name":"불안정 협심증 (UA)","cat":CAT,"diff":"★★★☆☆",
 "kw":"안정시 흉통 / Crescendo / Troponin 정상 / ACS 스펙트럼 / 배제진단",
 "mentor":'"\'협심증인데 불안정\'이 무슨 뜻일까? 전에는 계단 오를 때만 아프던 게, 이제 가만히 있어도 아프면 곧 막힐 신호야. Troponin이 정상이라 안심되겠지만, 이건 NSTEMI 바로 전 단계야. 절대 가볍게 보지 마."',
 "sections":[
   {"title":"📖 정의","img":"🖼️ 안정형 vs 불안정형 협심증 — 플라크 안정도 비교","rows":[
     {"item":"한 줄 정의","content":"관상동맥 협착으로 인한 흉통이지만 심근 괴사는 아직 없는 상태(Troponin 정상). ACS의 가장 이른 단계.","basis":"플라크가 불안정해져 혈류가 위태롭지만 완전·부분 폐색에 이르진 않음 → 괴사 표지자(Troponin) 정상."},
     {"item":"불안정의 3유형","content":"① 안정시 발생 ② 새로 생긴 중증 협심증 ③ 점점 심해지는(crescendo) 협심증","basis":"기존 안정형과 달리 \'예측 불가\'·\'악화\' 패턴 → 곧 폐색될 수 있다는 경고."},
   ]},
   {"title":"🔬 병태생리","img":"","rows":[
     {"item":"기전","content":"플라크 파열·미란 → 일시적 혈전·혈관수축 → 간헐적 혈류 감소(괴사 직전)","basis":"혈전이 자연 용해되거나 혈관 수축이 풀리면 증상 완화 → 변동성. 다음엔 완전 폐색될 수 있음."},
   ]},
   {"title":"😣 증상","img":"","rows":[
     {"item":"증상","content":"안정시·경미한 활동에도 흉통, 지속시간 길어짐, NTG 반응 둔화","basis":"산소 공급-요구 불균형이 휴식 중에도 발생 → 협착이 심해졌다는 의미."},
   ]},
   {"title":"📈 검사","img":"🖼️ UA에서의 일과성 ST 하강/T 역전 ECG","rows":[
     {"item":"ECG","content":"흉통 시 일과성 ST 하강·T파 역전, 통증 멎으면 정상화 가능","basis":"허혈이 가역적 → 변화도 일과성. 통증 중 ECG가 진단에 유리."},
     {"item":"Serial Troponin","content":"0h·3~6h 모두 정상이어야 UA. 상승하면 NSTEMI로 재분류","basis":"UA와 NSTEMI 구분은 Troponin뿐 → 반드시 연속 측정으로 괴사 배제."},
   ]},
   {"title":"💊 치료 근거","img":"","rows":[
     {"item":"기본 처치","content":"모니터링, IV, 산소(필요시), Aspirin, NTG, β차단제","basis":"NSTEMI에 준해 항혈소판·항허혈로 진행 차단. ACS 스펙트럼으로 동일 선상 관리."},
     {"item":"입원 관찰","content":"흉통통제실/관찰병상, Serial Troponin·ECG, 위험도 평가","basis":"괴사로 진행하는지 감시. 고위험은 침습적 평가(관상동맥조영) 고려."},
   ]},
 ],
 "redflags":[
   "안정시 지속 흉통 + NTG 무반응 → 임박한 완전 폐색 신호, 즉시 보고",
   "Troponin 상승 전환 → 더 이상 UA 아님, NSTEMI로 격상 대응",
   "흉통 중 ECG 변화 포착이 중요 → 통증 시 즉시 12-lead",
   "\'협심증 약 받으러 왔어요\' 환자도 패턴 악화면 ACS로 접근",
 ],
 "memo":"Troponin 정상 = UA / 상승 = NSTEMI / 안정시·악화 패턴 = 폐색 임박 / Aspirin+NTG, Serial 추적",
})

CASES.append({
 "no":"04","name":"급성 심부전 / 폐부종","cat":CAT,"diff":"★★★★☆",
 "kw":"좌심실 부전 / 폐울혈 / 기좌호흡 / BNP / LMNOP / NIV·이뇨제",
 "mentor":'"숨차서 못 눕는 환자 오면 폐부종부터 떠올려. 앉으면 좀 낫고 누우면 더 힘들어하지? 그게 핵심 단서야. 일단 앉히고 산소 줘. 약 순서는 LMNOP로 외우면 편해. 단, 모르핀은 요즘 조심해서 써."',
 "sections":[
   {"title":"📖 정의","img":"🖼️ 폐포로 체액이 새어 나온 폐부종 단면도","rows":[
     {"item":"한 줄 정의","content":"심장 펌프 기능 저하로 폐·전신에 체액이 정체되는 상태. 급성은 주로 폐부종으로 발현.","basis":"좌심실이 못 짜냄 → 좌심방·폐정맥 압력↑ → 폐모세혈관에서 폐포로 체액 누출 → 가스교환 장애."},
     {"item":"좌심부전 vs 우심부전","content":"좌심부전: 폐울혈(호흡곤란·기좌호흡·crackle)\n우심부전: 전신울혈(경정맥 팽대·하지부종·간비대)","basis":"막힌 쪽 \'뒤\'로 피가 정체됨. 좌심 뒤=폐, 우심 뒤=전신정맥."},
   ]},
   {"title":"🔬 병태생리","img":"","rows":[
     {"item":"악순환","content":"심박출↓ → 신장 관류↓ → RAAS·교감신경 활성 → 수분저류·후부하↑ → 심장 부담 가중","basis":"보상기전이 오히려 전·후부하를 늘려 악화 → 이뇨제·혈관확장제로 이 고리를 끊는다."},
   ]},
   {"title":"😣 증상","img":"","rows":[
     {"item":"증상별 근거","content":"기좌호흡·발작야간호흡곤란: 누우면 정맥환류↑ → 폐울혈 악화\n분홍빛 거품 가래: 폐포 누출액+공기\n양폐 crackle: 폐포 내 액체\n불안·발한: 저산소+교감항진","basis":"체위에 따른 정맥환류 변화가 증상을 좌우 → 앉히면 전부하 감소로 완화."},
   ]},
   {"title":"📈 검사","img":"🖼️ 폐부종 흉부 X선 — 양측 음영·Kerley B선·심비대","rows":[
     {"item":"BNP/NT-proBNP","content":"상승 → 심부전 시사(호흡곤란 감별)","basis":"심실벽 신장 → BNP 분비. 호흡곤란이 심장성인지 폐성인지 구분에 유용."},
     {"item":"흉부 X선","content":"폐문부 음영, Kerley B선, 흉수, 심비대","basis":"폐울혈·체액 정체의 직접 증거."},
     {"item":"기타","content":"ECG(허혈·부정맥 유발인자), Troponin, 전해질, 동맥혈가스","basis":"유발 원인(허혈·부정맥) 찾고 산소화·산염기 평가."},
   ]},
   {"title":"💊 치료 근거","img":"🖼️ LMNOP 처치 흐름 + NIV 적용 그림","rows":[
     {"item":"체위·산소","content":"상체 거상(앉히기) + 산소/NIV(CPAP/BiPAP)","basis":"앉히면 전부하↓. NIV 양압은 폐포 액체 밀어내고 호흡일 감소 → 삽관 회피."},
     {"item":"이뇨제(L)","content":"Furosemide IV","basis":"체액 배출로 전부하↓·폐울혈 완화. 정맥확장 효과도 일부."},
     {"item":"혈관확장(N)","content":"Nitroglycerin(SBP 적정 시)","basis":"정맥·동맥 확장 → 전·후부하↓ → 심장 부담 즉각 감소."},
     {"item":"모르핀(M)","content":"통증·불안 시 신중히(호흡억제 주의)","basis":"불안·전부하 감소 효과 있으나 호흡억제·예후 논란 → 일상 권장 아님, 선택적."},
     {"item":"원인 교정","content":"허혈·부정맥·고혈압·판막 등 유발인자 치료","basis":"근본 원인 제거 없으면 재발. 급성 악화의 방아쇠를 찾는다."},
   ]},
 ],
 "redflags":[
   "분홍빛 거품 가래 + 양폐 crackle → 전형적 급성 폐부종, 즉시 산소·NIV",
   "SBP<90 + 말초 차가움 + 의식저하 → 심인성 쇼크, 강심제·중환자 보고",
   "NIV에도 산소포화 안 오르고 지치는 호흡 → 기관삽관 준비",
   "급성 악화엔 항상 유발인자(허혈·부정맥) 동반 → ECG·Troponin 확인",
 ],
 "memo":"기좌호흡=폐부종 단서 / 앉히고 산소·NIV / LMNOP(이뇨·혈관확장 핵심, 모르핀 신중) / 원인 교정",
})

CASES.append({
 "no":"05","name":"고혈압성 위기","cat":CAT,"diff":"★★★☆☆",
 "kw":"Urgency vs Emergency / 표적장기손상 / MAP 25% / IV 강압제 / 급강압 금기",
 "mentor":'"혈압 230이라고 무조건 확 떨어뜨리면 안 돼. 핵심은 \'장기 손상이 있냐\'야. 손상 없으면(urgency) 천천히, 손상 있으면(emergency) IV로. 그래도 첫 1시간에 25% 이상 떨어뜨리면 뇌·콩팥이 오히려 허혈돼. 천천히가 안전이야."',
 "sections":[
   {"title":"📖 정의","img":"🖼️ Urgency vs Emergency 분류 도식 (표적장기손상 유무)","rows":[
     {"item":"한 줄 정의","content":"보통 SBP>180 또는 DBP>120의 심한 고혈압. 표적장기손상 동반 여부로 둘로 나뉜다.","basis":"숫자 자체보다 \'장기 손상\'이 응급도를 결정. 같은 혈압이라도 손상 유무로 처치가 완전히 달라짐."},
     {"item":"Urgency vs Emergency","content":"Urgency: 표적장기손상 없음 → 경구약·외래, 수 시간~일 단위 강압\nEmergency: 손상 있음(뇌·심장·신장·대동맥·눈) → IV, 즉시","basis":"손상 없으면 급강압이 더 위험. 손상 있으면 진행 막기 위해 즉시 조절."},
   ]},
   {"title":"🔬 병태생리","img":"","rows":[
     {"item":"기전","content":"극심한 혈압 → 혈관내피 손상·섬유소양 괴사 → 장기 허혈·출혈","basis":"자동조절(autoregulation) 한계 초과 → 미세혈관 손상 → 뇌증·신손상·심부전 등."},
   ]},
   {"title":"😣 증상 (표적장기손상 신호)","img":"","rows":[
     {"item":"장기별 신호","content":"뇌: 심한 두통·시야장애·의식저하·국소신경징후(고혈압뇌증·뇌졸중)\n심장: 흉통·호흡곤란(ACS·폐부종)\n신장: 핍뇨·Cr↑\n대동맥: 찢어지는 등 통증(박리)\n눈: 시력저하(망막출혈·유두부종)","basis":"어느 장기가 손상됐는지가 강압 목표·약물 선택을 좌우 → 증상으로 손상 장기 추적."},
   ]},
   {"title":"📈 검사","img":"","rows":[
     {"item":"필수 평가","content":"양팔 혈압, ECG, Troponin, BUN/Cr·전해질, 소변검사, 흉부X선, 신경학적 진찰; 필요시 뇌CT·대동맥CT, 안저","basis":"표적장기손상 스크리닝. 양팔 혈압 차이는 대동맥 박리 단서."},
   ]},
   {"title":"💊 치료 근거","img":"🖼️ MAP 시간별 강압 목표 그래프 (첫1h 25%)","rows":[
     {"item":"강압 목표","content":"Emergency: 첫 1시간 MAP 약 25%↓ → 이후 2~6h에 160/100 → 24~48h 정상화","basis":"뇌·신장 자동조절이 고혈압에 적응돼 있어 급강압 시 분수령 허혈. 단계적 강압이 안전."},
     {"item":"IV 약물","content":"Labetalol, Nicardipine, Nitroprusside, Esmolol 등 적정주입","basis":"적정 가능한 IV제로 부드럽게 조절. 질환별로 선택(아래)."},
     {"item":"질환별 선택","content":"대동맥박리: β차단제 먼저(HR<60, SBP100~120)\n폐부종: NTG+이뇨제\n뇌졸중: 기준 따라 신중 강압","basis":"박리는 맥압·shear stress 줄이려 β차단 우선. 질환마다 목표·약이 다름."},
     {"item":"Urgency","content":"경구약 재개·조정, 외래 추적","basis":"손상 없으므로 급할 필요 없음. 급강압이 오히려 위험."},
   ]},
 ],
 "redflags":[
   "찢어지는 흉통/등통 + 양팔 혈압 차이 → 대동맥 박리, β차단제 먼저",
   "심한 두통+의식저하+국소신경징후 → 고혈압뇌증/뇌졸중, 뇌영상",
   "첫 1시간 MAP 25% 초과 급강압 금지 → 장기 허혈 유발",
   "표적장기손상 없으면(urgency) IV 급강압 하지 말 것",
 ],
 "memo":"손상 유무가 핵심(Urgency vs Emergency) / 첫1h MAP 25%만↓ / 박리=β차단 먼저 / 양팔 혈압 필수",
})

CASES.append({
 "no":"07","name":"심실빈맥 (VT)","cat":CAT,"diff":"★★★★☆",
 "kw":"Wide QRS / 단형·다형 / 맥박유무 / 동기화 제세동 / Amiodarone·Mg",
 "mentor":'"넓은 QRS 빈맥 보면 일단 VT로 생각해. 맥 있나 없나, 안정하냐 불안정하냐 — 이 두 가지로 처치가 갈려. 맥 없으면 바로 제세동·CPR, 맥 있는데 불안정하면 동기화 심율동전환. 헷갈리면 VT로 대응하는 게 안전해."',
 "sections":[
   {"title":"📖 정의","img":"🖼️ 단형성 VT ECG — 넓고 규칙적인 QRS 연속","rows":[
     {"item":"한 줄 정의","content":"심실에서 시작된 빈맥(보통 HR>100, 넓은 QRS≥120ms)이 3개 이상 연속되는 부정맥.","basis":"정상 전도(His-Purkinje)를 안 거치고 심실 근육으로 퍼져 QRS가 넓고 변형됨."},
     {"item":"분류","content":"단형성(monomorphic): 모양 일정\n다형성(polymorphic): 모양 변동, Torsades=QT연장 동반\n지속(>30s) vs 비지속","basis":"다형성·Torsades는 원인·치료가 다름(Mg). 지속·혈역학 불안정일수록 위험."},
   ]},
   {"title":"🔬 병태생리","img":"","rows":[
     {"item":"기전·유발","content":"허혈·심근반흔의 회귀회로, 전해질이상(K·Mg), QT연장약물","basis":"VT 대부분은 기질적 심질환(허혈·심근증) 바탕. 빠른 심실수축 → 심실 충만 부족 → 심박출 급감."},
   ]},
   {"title":"😣 증상","img":"","rows":[
     {"item":"증상","content":"두근거림·실신·흉통·호흡곤란, 심하면 무맥성 심정지","basis":"심박출 급감으로 뇌·관상 관류 저하. 무맥 VT는 즉시 심정지 알고리즘."},
   ]},
   {"title":"📈 검사","img":"🖼️ VT vs SVT(편위전도) 감별 포인트 도식","rows":[
     {"item":"ECG","content":"넓은 QRS 규칙 빈맥, 방실해리·융합/포획박동 → VT 강력 시사","basis":"방실해리(P·QRS 독립)는 VT의 결정적 단서. 애매하면 VT로 간주가 안전."},
     {"item":"전해질·기타","content":"K·Mg·Ca, Troponin, 약물력, QT 측정","basis":"교정 가능한 원인(저K·저Mg·QT연장) 찾기. Torsades면 Mg가 1차."},
   ]},
   {"title":"💊 치료 근거","img":"🖼️ 맥박·안정성 따른 VT 처치 알고리즘","rows":[
     {"item":"무맥성 VT","content":"즉시 제세동(비동기) + CPR + Epinephrine + Amiodarone","basis":"VF와 동일한 심정지 처치. 빠른 제세동이 생존 결정."},
     {"item":"맥박O·불안정","content":"동기화 심율동전환(synchronized cardioversion)","basis":"QRS에 동기화해 충격 → R-on-T로 VF 유발 방지. 불안정(저혈압·의식저하·흉통)이면 즉시."},
     {"item":"맥박O·안정","content":"Amiodarone IV (또는 Procainamide), 원인 교정","basis":"약물로 율동 전환 시도. 시간 여유 있으므로 항부정맥제 우선."},
     {"item":"다형성/Torsades","content":"MgSO₄ IV, 유발약 중단, K 교정, 필요시 제세동","basis":"Torsades는 QT연장 기반 → Mg가 핵심. 무맥이면 비동기 제세동."},
   ]},
 ],
 "redflags":[
   "넓은 QRS 빈맥은 일단 VT로 대응 (SVT로 단정 금지)",
   "무맥성 VT → 즉시 제세동·CPR (약 기다리지 말 것)",
   "맥박 있어도 불안정(저혈압·의식저하·흉통) → 동기화 심율동전환",
   "다형성/Torsades → MgSO₄, QT 연장 약물 즉시 중단",
 ],
 "memo":"넓은QRS=VT간주 / 무맥→제세동+CPR / 맥O 불안정→동기화전환 / 안정→Amiodarone / Torsades→Mg",
})

CASES.append({
 "no":"08","name":"심방세동 (AF)","cat":CAT,"diff":"★★★☆☆",
 "kw":"불규칙 불규칙 / P파 소실 / 율박수조절 / 항응고 CHA2DS2-VASc / 색전",
 "mentor":'"맥이 \'불규칙하게 불규칙\'하면 AF야. ECG에 P파가 안 보이지. 두 가지만 기억해 — 빠르면 심박수 떨어뜨리고(rate control), 피떡 안 생기게 항응고. 그리고 48시간 넘은 AF를 함부로 정상으로 돌리면(cardioversion) 굳어있던 혈전이 튀어 뇌졸중 와."',
 "sections":[
   {"title":"📖 정의","img":"🖼️ AF ECG — P파 소실·불규칙 R-R 간격","rows":[
     {"item":"한 줄 정의","content":"심방이 무질서하게 떨려(350~600회/분) 효과적 수축을 못 하는 부정맥. 맥박이 불규칙하게 불규칙.","basis":"여러 회귀회로가 심방에서 난립 → 방실결절이 일부만 통과시켜 심실반응이 불규칙."},
     {"item":"위험","content":"① 빠른 심실반응(RVR) → 심박출↓·심부전\n② 심방 정체 → 좌심방이 혈전 → 색전성 뇌졸중","basis":"수축 못 하는 심방(특히 좌심방이)에 피가 고여 혈전 → 떨어져 나가면 뇌·전신 색전."},
   ]},
   {"title":"🔬 병태생리","img":"","rows":[
     {"item":"유발인자","content":"고혈압·판막질환·갑상선기능항진·음주·심부전·폐질환·패혈증","basis":"심방 확장·자극이 회귀회로 형성. 급성 유발인자(갑상선·감염) 교정이 중요."},
   ]},
   {"title":"😣 증상","img":"","rows":[
     {"item":"증상","content":"두근거림·불규칙 맥, 어지럼·피로·호흡곤란, 무증상도 흔함","basis":"심박출 변동·빈맥으로 인한 증상. 무증상 AF도 색전 위험은 동일 → 우연 발견도 평가."},
   ]},
   {"title":"📈 검사","img":"","rows":[
     {"item":"ECG","content":"P파 소실, 불규칙한 R-R, f파(잔물결)","basis":"무질서한 심방 활동 → 뚜렷한 P 대신 잔물결. R-R 완전 불규칙이 특징."},
     {"item":"평가","content":"갑상선기능, 전해질, 심초음파(혈전·판막·기능), 발생 시점 확인","basis":"유발 원인·구조 평가 + 발생<48h 여부(cardioversion 안전성 판단)."},
   ]},
   {"title":"💊 치료 근거","img":"🖼️ Rate vs Rhythm + 항응고 의사결정 흐름","rows":[
     {"item":"불안정 시","content":"즉시 동기화 심율동전환","basis":"저혈압·허혈·심부전 등 혈역학 붕괴면 율동 회복이 우선."},
     {"item":"심박수 조절","content":"β차단제 또는 비DHP CCB(Diltiazem/Verapamil), 일부 Digoxin","basis":"방실결절 전도 늦춰 심실반응↓ → 심박출 회복·증상 완화. 대개 1차 전략."},
     {"item":"율동 조절","content":"항부정맥제/전기적 심율동전환(적응증 시)","basis":"증상 지속·젊은 환자 등에서 정상동율 회복 시도."},
     {"item":"항응고","content":"CHA₂DS₂-VASc로 위험도 평가 → 항응고제(DOAC/Warfarin)","basis":"색전성 뇌졸중 예방이 AF 관리의 핵심. 점수 기반으로 출혈위험과 저울질."},
     {"item":"Cardioversion 주의","content":"AF>48h(또는 불명)면 전처치 항응고 3주 또는 TEE로 혈전 배제 후 시행","basis":"고여 있던 혈전이 율동 회복 시 떨어져 색전 → 사전 항응고·TEE 필수."},
   ]},
 ],
 "redflags":[
   "혈역학 불안정 RVR → 즉시 동기화 심율동전환",
   "AF>48h를 항응고 없이 cardioversion → 색전성 뇌졸중 위험",
   "갑상선기능항진·패혈증 등 유발인자 동반 확인 (원인 교정)",
   "WPW 동반 AF에 방실결절 차단제(Digoxin/Verapamil) 금기 → 심실세동 유발",
 ],
 "memo":"불규칙불규칙·P소실 / 빠르면 rate control / CHA₂DS₂-VASc로 항응고 / >48h는 항응고 후 전환",
})

CASES.append({
 "no":"09","name":"완전방실차단 (3rd AVB)","cat":CAT,"diff":"★★★★☆",
 "kw":"방실해리 / P·QRS 독립 / 서맥 / 이탈박동 / 경피 페이싱",
 "mentor":'"3도 차단은 심방이랑 심실이 따로 노는 거야. P파랑 QRS가 각자 자기 박자로 가지. 심박수 30~40으로 뚝 떨어지면 아트로핀 줘도 잘 안 들어, 특히 QRS 넓은 이탈박동이면. 그땐 망설이지 말고 경피 페이싱 준비해."',
 "sections":[
   {"title":"📖 정의","img":"🖼️ 3도 방실차단 ECG — P와 QRS가 무관하게 진행","rows":[
     {"item":"한 줄 정의","content":"방실결절/전도계가 완전히 차단되어 심방 신호가 심실로 전혀 전달되지 않는 상태.","basis":"심방(P)과 심실(QRS)이 각자 독립 박동(방실해리) → 심실은 느린 이탈박동에 의존."},
     {"item":"이탈박동 위치","content":"접합부 이탈: QRS 좁음, 40~60회(비교적 안정)\n심실 이탈: QRS 넓음, 20~40회(불안정·불신뢰)","basis":"차단 부위가 낮을수록 이탈박동이 느리고 불안정 → 넓은 QRS 이탈은 응급."},
   ]},
   {"title":"🔬 병태생리","img":"","rows":[
     {"item":"원인","content":"하벽 심근경색(RCA), 전도계 퇴행성 섬유화, 약물(β차단제·CCB·Digoxin), 고칼륨혈증","basis":"교정 가능한 원인(약물·고K·허혈) 동반 흔함 → 원인 찾으면 가역적일 수 있음."},
   ]},
   {"title":"😣 증상","img":"","rows":[
     {"item":"증상","content":"실신·전실신, 어지럼·피로, 운동 못 견딤, 흉통·심부전, Adams-Stokes 발작","basis":"심박출 급감으로 뇌관류 저하 → 실신. 서맥이 심하면 심정지로 진행."},
   ]},
   {"title":"📈 검사","img":"","rows":[
     {"item":"ECG","content":"P-P 규칙·R-R 규칙이지만 P와 QRS 무관(PR 일정치 않음), 심실수 < 심방수","basis":"방실해리가 핵심 소견. P가 QRS보다 많고 둘이 독립적이면 3도."},
     {"item":"원인 평가","content":"전해질(K), 약물력, 심근효소/ECG(하벽 MI), 디곡신 농도","basis":"가역 원인 교정이 우선. 하벽 MI 동반 시 재관류로 회복 가능."},
   ]},
   {"title":"💊 치료 근거","img":"🖼️ 증상성 서맥 알고리즘 — 아트로핀→페이싱","rows":[
     {"item":"즉시 처치","content":"모니터·IV·산소, 증상성 서맥이면 Atropine 시도","basis":"Atropine은 미주신경 차단으로 심박↑ 시도. 단 결절하 차단(넓은 QRS)엔 효과 적음."},
     {"item":"경피 페이싱(TCP)","content":"Atropine 무효/넓은QRS 이탈/불안정 → 즉시 경피 페이싱","basis":"외부 전기자극으로 심실 박동 확보. 3도 차단의 핵심 응급 처치."},
     {"item":"약물 보조","content":"Dopamine 또는 Epinephrine 주입(페이싱 준비 중 가교)","basis":"심박수·관류 유지용 임시 수단. 경정맥 영구 페이싱으로 가는 다리."},
     {"item":"근본 치료","content":"가역 원인 교정 후에도 지속 시 경정맥 임시→영구 인공심박동기","basis":"전도계 비가역 손상이면 페이스메이커가 확정 치료."},
   ]},
 ],
 "redflags":[
   "넓은 QRS·HR 20~40 이탈박동 → 불안정·심정지 임박, 즉시 경피 페이싱",
   "Atropine에 반응 없는 서맥 → 페이싱으로 바로 전환 (반복 투여로 시간 낭비 금지)",
   "하벽 MI 동반 3도 차단 → 재관류로 가역 가능, 즉시 보고",
   "고칼륨혈증·약물(Digoxin·β차단제) 원인 → 교정하면 회복 가능",
 ],
 "memo":"P·QRS 독립(방실해리) / 넓은QRS 이탈=위험 / Atropine 무효 많음 → 경피 페이싱 / 가역원인 교정",
})

CASES.append({
 "no":"10","name":"심낭염 / 심낭압전","cat":CAT,"diff":"★★★★☆",
 "kw":"흉막성 흉통 / 광범위 ST상승·PR하강 / Beck 3징 / 기이맥 / 심낭천자",
 "mentor":'"심낭염 흉통은 AMI랑 헷갈려. 근데 \'앉아서 앞으로 숙이면 편하고 누우면 아프다\'·\'숨 들이쉬면 아프다\'면 심낭염 쪽이야. ECG도 광범위하게 ST가 올라가지(한 혈관 영역 아님). 진짜 무서운 건 압전 — 혈압 떨어지고 목정맥 튀어나오고 심음 멀어지면 바로 심낭천자야."',
 "sections":[
   {"title":"📖 정의","img":"🖼️ 심낭 구조 + 삼출액 고인 심낭압전 단면","rows":[
     {"item":"한 줄 정의","content":"심낭(심장 주머니)의 염증이 심낭염, 삼출액이 차서 심장을 눌러 채워지지 못하게 하는 게 심낭압전.","basis":"심낭은 신축성이 적음 → 액체가 빠르게 차면 적은 양으로도 심실 충만 방해(압전)."},
     {"item":"원인","content":"바이러스·특발성(흔함), 결핵, 요독증, 외상, MI 후, 악성, 자가면역","basis":"원인에 따라 경과·치료가 다름. 급성 압전은 외상·박리·시술 합병 등 빠르게 차는 경우."},
   ]},
   {"title":"🔬 병태생리","img":"","rows":[
     {"item":"압전 기전","content":"심낭압↑ → 우심실 이완기 충만 방해 → 정맥환류 정체·심박출↓ → 쇼크","basis":"\"못 채워지면 못 짜낸다\" → 전부하 의존. 그래서 수액으로 일시 보조, 근본은 액체 빼기."},
   ]},
   {"title":"😣 증상","img":"","rows":[
     {"item":"심낭염","content":"흉막성 흉통(흡기·기침 시 악화), 앉아 앞으로 숙이면 완화, 심막마찰음","basis":"염증 심막이 호흡·체위로 자극됨 → AMI(체위 무관)와 구분되는 단서."},
     {"item":"압전(Beck 3징)","content":"저혈압 + 경정맥 팽대 + 심음 감약, 기이맥(흡기 시 SBP>10mmHg 하강), 빈맥","basis":"충만 장애로 혈압↓·정맥울혈↑, 액체가 심음 차단. 기이맥은 압전의 특징적 소견."},
   ]},
   {"title":"📈 검사","img":"🖼️ 심낭염 ECG — 광범위 ST상승·PR하강 / 압전 심초음파","rows":[
     {"item":"ECG","content":"심낭염: 광범위(여러 영역) 오목한 ST 상승 + PR 하강\n압전: 저전압·전기적 교대(electrical alternans)","basis":"한 관상동맥 영역에 국한된 STEMI와 달리 \'전반적\'. 전기적 교대는 심장이 액체 속에서 흔들리는 소견."},
     {"item":"영상","content":"심초음파(삼출·이완기 우심 허탈), 흉부X선(물주머니 심장)","basis":"압전 진단·심낭천자 유도의 핵심. 우심 허탈은 압전의 직접 증거."},
   ]},
   {"title":"💊 치료 근거","img":"🖼️ 심낭천자 시술 모식도","rows":[
     {"item":"심낭염","content":"NSAID + Colchicine, 원인 치료, 휴식","basis":"항염증으로 통증·재발 감소. Colchicine이 재발 예방에 근거."},
     {"item":"압전 — 응급","content":"즉시 심낭천자(pericardiocentesis)로 액체 제거","basis":"눌림 해소가 유일한 즉효 치료. 지체 시 무맥성 전기활동(PEA) 심정지."},
     {"item":"가교 처치","content":"천자 전까지 수액(전부하 유지), 산소, 모니터","basis":"전부하 의존 상태라 수액이 일시적 충만 보조. 단 근본 해결은 액체 제거."},
     {"item":"피해야 할 것","content":"과도한 이뇨·혈관확장(전부하↓) 금기","basis":"전부하 떨어뜨리면 충만 더 악화 → 압전에선 혈압 더 무너짐."},
   ]},
 ],
 "redflags":[
   "Beck 3징(저혈압·경정맥팽대·심음감약) + 기이맥 → 심낭압전, 즉시 심낭천자 준비",
   "압전 의심에 이뇨제·혈관확장제 금기 (전부하 의존)",
   "광범위 ST상승+PR하강은 심낭염 — STEMI와 감별(체위·흉막성 통증)",
   "외상·대동맥박리·시술 후 급성 압전 → 빠르게 진행, 즉시 보고",
 ],
 "memo":"앉아 숙이면 편함=심낭염 / 광범위 ST↑·PR↓ / Beck3징+기이맥=압전 / 압전→심낭천자, 이뇨 금기",
})

CASES.append({
 "no":"11","name":"대동맥 박리","cat":CAT,"diff":"★★★★★",
 "kw":"찢어지는 통증 / 양팔 혈압차 / 종격동 확대 / CT 혈관조영 / Stanford A·B",
 "mentor":'"\'칼로 찢는 것 같다\'·\'등으로 뻗친다\'는 흉통이면 박리를 의심해. 양팔 혈압 꼭 재 — 차이 나면 강한 단서야. 여기서 제일 중요한 건, 혈압·심박수를 먼저 낮춰서 찢기는 힘을 줄이는 거야. 그냥 혈관확장제만 주면 반사빈맥으로 더 찢어져. β차단제 먼저!"',
 "sections":[
   {"title":"📖 정의","img":"🖼️ 대동맥 내막 파열 → 가성내강 형성 단면도","rows":[
     {"item":"한 줄 정의","content":"대동맥 내막이 찢어져 혈액이 중막 사이로 파고들어 \'가성내강\'을 만드는 치명적 응급.","basis":"고압 혈류가 내막 파열부로 들어가 벽을 갈라놓음 → 분지혈관 폐색·파열·압전으로 사망 위험."},
     {"item":"Stanford 분류","content":"A형: 상행대동맥 침범 → 응급 수술\nB형: 하행만 → 대개 약물치료","basis":"A형은 관상동맥·대동맥판·심낭으로 파급 → 압전·MI·판막부전 위험이 커 수술적."},
   ]},
   {"title":"🔬 병태생리","img":"","rows":[
     {"item":"위험인자","content":"고혈압(최대), 결합조직질환(Marfan), 이엽성 대동맥판, 외상, 임신","basis":"벽 스트레스·구조 취약 → 내막 파열. 혈압·맥압(shear stress)이 진행을 가속."},
   ]},
   {"title":"😣 증상","img":"","rows":[
     {"item":"증상","content":"갑작스런 찢어지는 흉통/등통(이동성), 양팔 혈압·맥박 차이, 실신, 분지 침범 시 뇌졸중·하지허혈·핍뇨","basis":"박리가 분지혈관을 막으며 다양한 허혈 증상. 양팔 혈압차는 쇄골하동맥 침범 단서."},
   ]},
   {"title":"📈 검사","img":"🖼️ 흉부X선 종격동 확대 / CT 혈관조영 가성내강","rows":[
     {"item":"흉부 X선","content":"종격동 확대, 대동맥 윤곽 이상","basis":"빠른 선별 단서. 정상이라도 박리 배제는 못 함."},
     {"item":"CT 혈관조영","content":"확진 검사 — 내막 피판·가성내강·침범 범위","basis":"분류(A/B)·수술 결정의 기준. 혈역학 안정 시 1차 확진."},
     {"item":"기타","content":"양팔 혈압, ECG(MI 감별), 심초음파(압전·판막), D-dimer","basis":"AMI·압전 동반 평가. 양팔 혈압차는 침상에서 즉시 가능한 단서."},
   ]},
   {"title":"💊 치료 근거","img":"🖼️ 박리 강압 전략 — β차단 먼저, 목표 SBP100~120·HR<60","rows":[
     {"item":"1차: β차단제","content":"Esmolol/Labetalol로 HR<60, SBP 100~120 먼저","basis":"심박수·수축력↓ → 대동맥 벽 shear stress(dP/dt)↓ → 박리 진행 억제. 반드시 혈관확장제보다 먼저."},
     {"item":"2차: 혈관확장","content":"β차단 후에도 혈압 높으면 Nicardipine/Nitroprusside 추가","basis":"먼저 혈관만 확장하면 반사성 빈맥 → shear↑ → 더 찢어짐. 순서가 생명."},
     {"item":"통증 조절","content":"Morphine 등으로 통증·교감항진 완화","basis":"통증→교감↑→혈압·맥박↑→박리 악화 고리 차단."},
     {"item":"확정 치료","content":"A형: 즉시 흉부외과 수술 / B형: 약물·합병증 시 중재","basis":"A형은 파열·압전 위험으로 수술이 원칙. B형은 약물조절 우선."},
   ]},
 ],
 "redflags":[
   "찢어지는 흉통/등통 + 양팔 혈압차 → 대동맥 박리 강력 의심",
   "혈관확장제만 단독 투여 금지 → 반사빈맥으로 박리 악화 (β차단 먼저)",
   "A형(상행 침범) → 심낭압전·MI 동반 가능, 즉시 흉부외과",
   "박리를 ACS로 오인해 항응고·혈전용해 → 치명적 출혈 위험",
 ],
 "memo":"찢어지는 통증·양팔 혈압차 / CT혈관조영 확진 / β차단 먼저→SBP100~120·HR<60 / A형=수술",
})

# ════════════════════════════════════════════════════════════════
def main():
    import shutil
    shutil.copyfile(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    completed = {}
    for case in CASES:
        sn = render_case(wb, case)
        completed[case["no"]] = sn
        print(f"  ✅ {case['no']} {case['name']} → '{sn}'")
    update_list(wb, completed)
    wb.save(DST)
    print(f"\n저장 완료: {DST}")
    print(f"추가된 케이스: {len(CASES)}개")
    print(f"전체 시트: {wb.sheetnames}")

if __name__ == "__main__":
    main()

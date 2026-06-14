#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
응급실 교육케이스 — 공용 렌더링 엔진
내과 생성기(_gen_cases_full.py)와 동일한 8단계·이론적근거 레이아웃.
새 워크북을 처음부터 생성(목록 시트 자동 빌드). 그림(이미지) 생략.
"""
import os, math, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from _paths import FINAL_DIR  # 결과물은 01 폴더로 저장

SEC_COLORS = ["1565C0", "1B5E20", "E65100", "6A1B9A", "B71C1C", "00838F"]
REDFLAG_HDR = "B71C1C"; REDFLAG_BODY = "FFEBEE"
MEMO_FILL = "1A237E"; HEADER_FILL = "1A2A3A"; MENTOR_FILL = "FFFDE7"
LIST_TITLE_FILL = "1A2A3A"; LIST_HDR_FILL = "37474F"

# 부서 카테고리 색상 (파일별로 필요한 것만 사용)
CAT_COLOR = {
    "🔪 외과/외상": "37474F", "🔥 화상": "BF360C", "🦴 정형외과": "4E342E",
    "🫁 흉부외과": "01579B", "🤰 산과/부인과": "880E4F", "👶 소아과": "00838F",
    "🧠 정신과": "4527A0", "🧠 신경외과": "283593",
}

def F(rgb): return PatternFill("solid", fgColor=rgb)
def white(sz, b=True): return Font(name="맑은 고딕", size=sz, bold=b, color="FFFFFF")
def dark(sz, b=False, color="212121"): return Font(name="맑은 고딕", size=sz, bold=b, color=color)
AL_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
AL_LT = Alignment(horizontal="left", vertical="top", wrap_text=True)
AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)

def S(title, rows): return {"title": title, "rows": rows}
def R(item, content, basis): return {"item": item, "content": content, "basis": basis}

def est_lines(text, width):
    if not text: return 1
    n = 0
    for ln in str(text).split("\n"):
        n += max(1, math.ceil(len(ln) / width))
    return n

def _safe(name):
    return name[:31].replace("/", "·").replace("\\", "·").replace("?", "").replace("*", "").replace("[", "(").replace("]", ")").replace(":", "·")

def render_case(wb, case):
    no, name, cat = case["no"], case["name"], case["cat"]
    diff, kw, mentor = case["diff"], case["kw"], case["mentor"]
    sheet_name = _safe(f"{no}_{name}")
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    catfill = CAT_COLOR.get(cat, "455A64")

    for col, w in zip("ABCDE", [5, 22, 40, 40, 2]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:D1")
    ws["A1"] = f"CASE {no}  |  {name}"
    ws["A1"].fill = F(catfill); ws["A1"].font = white(14, True); ws["A1"].alignment = AL_L
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A2:D2")
    ws["A2"] = f"{cat}  |  난이도: {diff}  |  {kw}"
    ws["A2"].fill = F(catfill); ws["A2"].font = white(9, False); ws["A2"].alignment = AL_L
    ws.row_dimensions[2].height = 18
    ws.merge_cells("A3:D3")
    ws["A3"] = f"💬 선배 말씀 (프셉마음)\n{mentor}"
    ws["A3"].fill = F(MENTOR_FILL); ws["A3"].font = dark(9.5); ws["A3"].alignment = AL_LT
    ws.row_dimensions[3].height = 58
    for col, txt in zip("ABCD", ["섹션", "항목", "내용", "이론적 근거 (왜?)"]):
        c = ws[f"{col}4"]; c.value = txt; c.fill = F(HEADER_FILL); c.font = white(10, True); c.alignment = AL_C
    ws.row_dimensions[4].height = 22

    r = 5
    for si, sec in enumerate(case["sections"]):
        color = SEC_COLORS[si % len(SEC_COLORS)]
        ws.merge_cells(f"A{r}:D{r}")
        ws[f"A{r}"] = f"  {sec['title']}"
        ws[f"A{r}"].fill = F(color); ws[f"A{r}"].font = white(10, True); ws[f"A{r}"].alignment = AL_L
        ws.row_dimensions[r].height = 22
        r += 1
        rows = sec["rows"]
        for ri, row in enumerate(rows):
            ws[f"B{r}"] = row.get("item", ""); ws[f"B{r}"].font = dark(9.5, True); ws[f"B{r}"].alignment = AL_LT
            ws[f"C{r}"] = row.get("content", ""); ws[f"C{r}"].font = dark(9.5); ws[f"C{r}"].alignment = AL_LT
            ws[f"D{r}"] = row.get("basis", ""); ws[f"D{r}"].font = dark(9.0); ws[f"D{r}"].alignment = AL_LT
            lines = max(est_lines(row.get("content",""), 38), est_lines(row.get("basis",""), 40), est_lines(row.get("item",""), 11))
            ws.row_dimensions[r].height = min(max(lines * 15 + 6, 30), 230)
            r += 1
            if ri < len(rows) - 1:
                ws.row_dimensions[r].height = 8; r += 1
        ws.row_dimensions[r].height = 8; r += 1

    ws.merge_cells(f"A{r}:D{r}")
    ws[f"A{r}"] = "  🔴 Red Flag — 절대 놓치면 안 되는 것"
    ws[f"A{r}"].fill = F(REDFLAG_HDR); ws[f"A{r}"].font = white(10, True); ws[f"A{r}"].alignment = AL_L
    ws.row_dimensions[r].height = 22; r += 1
    for rf in case["redflags"]:
        ws.merge_cells(f"A{r}:D{r}")
        ws[f"A{r}"] = f"  ⚠️  {rf}"
        ws[f"A{r}"].fill = F(REDFLAG_BODY); ws[f"A{r}"].font = dark(9.5, True, REDFLAG_HDR); ws[f"A{r}"].alignment = AL_L
        ws.row_dimensions[r].height = max(est_lines(rf, 95) * 15 + 6, 20); r += 1
    ws.row_dimensions[r].height = 8; r += 1

    ws.merge_cells(f"A{r}:D{r}")
    ws[f"A{r}"] = f"  ⚡ 핵심 암기  |  {case['memo']}"
    ws[f"A{r}"].fill = F(MEMO_FILL); ws[f"A{r}"].font = white(9.5, True); ws[f"A{r}"].alignment = AL_L
    ws.row_dimensions[r].height = max(est_lines(case["memo"], 100) * 15 + 10, 28)
    return sheet_name

def build_list(wb, cases, title, subtitle):
    """📋 케이스 목록 시트를 맨 앞에 생성하고 케이스 행 채움."""
    ws = wb.create_sheet("📋 케이스 목록", 0)
    for col, w in zip("ABCDEFG", [8, 30, 18, 12, 50, 10, 28]):
        ws.column_dimensions[col].width = w
    ws.merge_cells("A1:G1")
    ws["A1"] = title; ws["A1"].fill = F(LIST_TITLE_FILL); ws["A1"].font = white(14, True); ws["A1"].alignment = AL_L
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A2:G2")
    ws["A2"] = subtitle; ws["A2"].fill = F(LIST_TITLE_FILL); ws["A2"].font = white(9, False); ws["A2"].alignment = AL_L
    ws.row_dimensions[2].height = 18
    for col, txt in zip("ABCDEFG", ["No", "병명/주제", "카테고리", "난이도", "키워드", "상태", "시트 이름"]):
        c = ws[f"{col}4"]; c.value = txt; c.fill = F(LIST_HDR_FILL); c.font = white(10, True); c.alignment = AL_C
    ws.row_dimensions[4].height = 22
    r = 5
    for case in cases:
        sn = _safe(f"{case['no']}_{case['name']}")
        vals = [case["no"], case["name"], case["cat"], case["diff"], case["kw"], "✅ 완성", sn]
        for ci, v in enumerate(vals):
            c = ws.cell(r, ci + 1); c.value = v
            c.font = dark(9.5, ci in (0, 5)); c.alignment = AL_L if ci in (1, 4, 6) else AL_C
            c.fill = F(CAT_COLOR.get(case["cat"], "ECEFF1") if False else "F5F5F5")
        ws.cell(r, 3).fill = F(CAT_COLOR.get(case["cat"], "90A4AE")); ws.cell(r, 3).font = white(9, True)
        ws.row_dimensions[r].height = max(est_lines(case["kw"], 50) * 14 + 6, 26)
        r += 1

def build_workbook(path, cases, title, subtitle):
    # 파일명만 주어지면 01 폴더로 자동 저장(어느 위치에서 실행해도 동일)
    if not os.path.isabs(path) and os.path.dirname(path) == "":
        path = os.path.join(FINAL_DIR, path)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 시트 제거
    build_list(wb, cases, title, subtitle)
    for case in cases:
        sn = render_case(wb, case)
        print(f"  ✅ {case['no']} {case['name']} → '{sn}'")
    wb.save(path)
    print(f"\n저장 완료: {path}")
    print(f"케이스 수: {len(cases)} / 전체 시트: {len(wb.sheetnames)}")


def append_list_rows(wb, cases):
    """기존 📋 케이스 목록 시트의 마지막 데이터 행 뒤에 케이스 행을 이어 붙인다."""
    ws = wb["📋 케이스 목록"]
    r = 5
    while ws.cell(r, 1).value not in (None, ""):
        r += 1
    for case in cases:
        sn = _safe(f"{case['no']}_{case['name']}")
        vals = [case["no"], case["name"], case["cat"], case["diff"], case["kw"], "✅ 완성", sn]
        for ci, v in enumerate(vals):
            c = ws.cell(r, ci + 1); c.value = v
            c.font = dark(9.5, ci in (0, 5)); c.alignment = AL_L if ci in (1, 4, 6) else AL_C
            c.fill = F("F5F5F5")
        ws.cell(r, 3).fill = F(CAT_COLOR.get(case["cat"], "90A4AE")); ws.cell(r, 3).font = white(9, True)
        ws.row_dimensions[r].height = max(est_lines(case["kw"], 50) * 14 + 6, 26)
        r += 1


def append_workbook(src_path, dst_path, cases, title=None, subtitle=None):
    """기존 워크북(src)에 케이스를 추가해 새 파일(dst)로 저장(원본 보존)."""
    wb = openpyxl.load_workbook(src_path)
    ws = wb["📋 케이스 목록"]
    if title:
        ws["A1"] = title
    if subtitle:
        ws["A2"] = subtitle
    append_list_rows(wb, cases)
    for case in cases:
        sn = render_case(wb, case)
        print(f"  ✅ {case['no']} {case['name']} → '{sn}'")
    wb.save(dst_path)
    print(f"\n저장 완료: {dst_path}")
    print(f"추가 케이스: {len(cases)} / 전체 시트: {len(wb.sheetnames)}")

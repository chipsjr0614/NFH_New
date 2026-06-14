#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
응급실 교육·시나리오 4주 운영 스케줄 — 구글시트용 엑셀 생성기.
간호사15(중 ICU파견5)+응급구조사5, 데이/이브닝 반반.
각자 이름 입력→자가학습·개인발표·혼성팀 시나리오·진행체크.
Drive 업로드 → Google 스프레드시트로 열기 → 공유(편집)하면 실시간 협업.
"""
import os, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "응급실_교육운영_스케줄_4주_구글시트용.xlsx")

NAVY="1A2A3A"; HDR="37474F"; BLUE="1565C0"; GREEN="1B5E20"; ORANGE="E65100"
PURPLE="4527A0"; RED="B71C1C"; LGRAY="F5F5F5"; YEL="FFF8E1"
def F(c): return PatternFill("solid", fgColor=c)
def W(sz,b=True): return Font(name="맑은 고딕", size=sz, bold=b, color="FFFFFF")
def D(sz,b=False,c="212121"): return Font(name="맑은 고딕", size=sz, bold=b, color=c)
AL=Alignment(horizontal="left", vertical="center", wrap_text=True)
AC=Alignment(horizontal="center", vertical="center", wrap_text=True)
thin=Side(style="thin", color="D5DBE1"); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)

wb=openpyxl.Workbook(); wb.remove(wb.active)
STATUS=['☐ 예정','▶ 진행','✅ 완료']

def title_row(ws, text, sub, ncol):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncol)
    c=ws.cell(1,1); c.value=text; c.fill=F(NAVY); c.font=W(14); c.alignment=AL
    ws.row_dimensions[1].height=30
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncol)
    c=ws.cell(2,1); c.value=sub; c.fill=F(NAVY); c.font=W(9,False); c.alignment=AL
    ws.row_dimensions[2].height=18

def header(ws, row, cols, fill=HDR):
    for i,t in enumerate(cols,1):
        c=ws.cell(row,i); c.value=t; c.fill=F(fill); c.font=W(10); c.alignment=AC; c.border=BORD
    ws.row_dimensions[row].height=24

def cell(ws,r,c,v,bold=False,fill=None,al=AL,color="212121",sz=10):
    x=ws.cell(r,c); x.value=v; x.font=D(sz,bold,color); x.alignment=al; x.border=BORD
    if fill: x.fill=F(fill)
    return x

# ════════════════════════════════════════════════════════════
# 1) 📋 사용안내
# ════════════════════════════════════════════════════════════
ws=wb.create_sheet("📋 사용안내")
for col,w in zip("AB",[3,110]): ws.column_dimensions[col].width=w
title_row(ws,"📋 응급실 교육·시나리오 4주 운영 — 사용 안내","간호사 15(ICU 파견 5 포함) + 응급구조사 5 · 데이/이브닝 · 자기주도 학습",2)
guide=[
 ("①","[공유 세팅] 이 파일을 Google Drive에 업로드 → 우클릭 → 'Google 스프레드시트로 열기' → 우상단 '공유'에서 팀원 전체에 '편집자' 권한."),
 ("②","[이름 입력] '👥 참가자' 탭에서 본인 칸에 이름을 적습니다. 근무(데이/이브닝)·ICU 파견 여부 확인."),
 ("③","[자가학습] '🗓️ 4주 스케줄'의 해당 주차 범위를 교육 프로그램(HTML)·케이스 엑셀로 미리 공부."),
 ("④","[개인 발표] '🎤 개인 발표 배정'에서 본인 주제·발표일 확인 → 발표자료 만들고 링크 칸에 붙임 → 끝나면 상태를 '✅ 완료'로."),
 ("⑤","[팀 시나리오] '👥 팀 시나리오 조'에서 본인 조·역할 확인 → 주차별 시나리오 롤플레이 진행 후 체크."),
 ("⑥","[진행 확인] '✅ 진행 체크보드'에서 전체 진행률을 한눈에. 각 칸은 드롭다운(☐예정/▶진행/✅완료)."),
 ("",""),
 ("규칙","• 데이/이브닝은 각 조끼리 같은 시간대에 모여 발표·시나리오 (조별 일정은 팀이 자율 조정)."),
 ("규칙","• ICU 파견 5명: 자가학습+퀴즈 필수, 발표·시나리오는 '가능한 근무일'에 부분 참여(불참 시 녹화 발표로 대체)."),
 ("규칙","• 발표자료는 5~10분 분량. 케이스의 '선배 말씀·Red Flag·핵심 암기'를 중심으로."),
 ("규칙","• 시나리오는 역할(리더/기도/약물/기록/구조사 처치)을 나눠 실제처럼. 끝나고 디브리핑 5분."),
 ("팁","• 셀에 메모/링크 자유롭게 추가. 변경 이력은 구글시트가 자동 저장."),
]
r=4
for tag,txt in guide:
    cell(ws,r,1,tag,bold=True,fill=YEL if tag in("규칙","팁") else None,al=AC,color=RED if tag in("규칙","팁") else "212121")
    cell(ws,r,2,txt); ws.row_dimensions[r].height=30 if txt else 8
    r+=1

# ════════════════════════════════════════════════════════════
# 2) 👥 참가자
# ════════════════════════════════════════════════════════════
ws=wb.create_sheet("👥 참가자")
for col,w in zip("ABCDEF",[6,12,16,12,14,30]): ws.column_dimensions[col].width=w
title_row(ws,"👥 참가자 명단 (총 20명)","본인 칸에 이름 입력 · 근무/파견은 실제에 맞게 수정",6)
header(ws,4,["번호","직역","이름(입력)","근무","ICU 파견","비고"])
# 제안 배정: 간호사 N1~N5 파견(이브닝/데이 무관·부분), N6~N10 데이, N11~N15 이브닝 / 구조사 E1~E3 데이, E4~E5 이브닝
rows=[]
for i in range(1,16):
    if i<=5: shift,disp="(파견)","✅ 파견"
    elif i<=10: shift,disp="데이","—"
    else: shift,disp="이브닝","—"
    rows.append((f"N{i:02d}","간호사","",shift,disp,"파견 5명은 가능일 부분참여" if i<=5 else ""))
for j in range(1,6):
    shift="데이" if j<=3 else "이브닝"
    rows.append((f"E{j:02d}","응급구조사","",shift,"—",""))
r=5
for no,role,name,shift,disp,memo in rows:
    fill=LGRAY if r%2 else "FFFFFF"
    cell(ws,r,1,no,bold=True,fill=fill,al=AC)
    cell(ws,r,2,role,fill=fill,al=AC,color=BLUE if role=="간호사" else GREEN)
    cell(ws,r,3,name,fill="FFFDE7",al=AC)  # 이름 입력칸 강조
    cell(ws,r,4,shift,fill=fill,al=AC)
    cell(ws,r,5,disp,fill=fill,al=AC,color=RED if "파견" in disp else "9AA7B5")
    cell(ws,r,6,memo,fill=fill)
    ws.row_dimensions[r].height=24; r+=1

# ════════════════════════════════════════════════════════════
# 3) 🗓️ 4주 스케줄 (마스터)
# ════════════════════════════════════════════════════════════
ws=wb.create_sheet("🗓️ 4주 스케줄")
for col,w in zip("ABCDE",[8,20,40,34,26]): ws.column_dimensions[col].width=w
title_row(ws,"🗓️ 4주 마스터 스케줄","주차별 주제·자가학습 범위·개인발표·팀 시나리오 (날짜는 팀이 입력)",5)
header(ws,4,["주차","주제","자가학습 범위(케이스)","개인 발표 주제","팀 시나리오"])
weeks=[
 ("1주차","심정지·치명적 부정맥","심장/순환기: 심정지(VF)·VT·심방세동·완전방실차단·아나필락시스",
  "①심실세동/심정지 ②심실빈맥(VT) ③아나필락시스 ④완전방실차단","심정지 발생 — CPR·제세동 팀대응"),
 ("2주차","흉통·호흡곤란","STEMI/NSTEMI·급성 폐색전증·천식/COPD 악화·급성 심부전/폐부종",
  "①STEMI ②급성 폐색전증(PE) ③천식·COPD 악화 ④급성 심부전·폐부종","흉통 환자 — STEMI 인지·D2B 90분"),
 ("3주차","의식저하·신경","허혈성/출혈성 뇌졸중·저혈당·경련/뇌전증 지속상태·뇌수막염",
  "①허혈성 뇌졸중 ②저혈당 ③뇌전증 지속상태 ④출혈성 뇌졸중","쓰러진 환자 — 뇌졸중 골든타임 vs 저혈당"),
 ("4주차","쇼크·패혈증·중독","패혈증/패혈성 쇼크·출혈성 쇼크·약물 중독·고칼륨혈증",
  "①패혈증·패혈성 쇼크 ②출혈성 쇼크·대량수혈 ③약물 과용량/중독 ④고칼륨혈증","패혈증 — Hour-1 번들 팀대응"),
]
r=5
wkfill={"1주차":RED,"2주차":ORANGE,"3주차":BLUE,"4주차":GREEN}
for wk,theme,study,pres,scen in weeks:
    cell(ws,r,1,wk,bold=True,fill=wkfill[wk],al=AC,color="FFFFFF")
    cell(ws,r,2,theme,bold=True)
    cell(ws,r,3,study); cell(ws,r,4,pres); cell(ws,r,5,scen)
    ws.row_dimensions[r].height=92; r+=1
cell(ws,r+1,2,"※ 데이조·이브닝조 각자 시간대에 모여 진행 · 날짜/시간은 조별로 아래 탭에서 조정",al=AL,color="8B98A8")

# ════════════════════════════════════════════════════════════
# 4) 🎤 개인 발표 배정
# ════════════════════════════════════════════════════════════
ws=wb.create_sheet("🎤 개인 발표 배정")
for col,w in zip("ABCDEFG",[8,12,30,12,16,30,12]): ws.column_dimensions[col].width=w
title_row(ws,"🎤 개인 발표 배정 (각자 1회 이상)","제안 담당(번호)은 바꿔도 됨 · 이름·발표일 입력 · 자료 링크 붙이기 · 끝나면 상태 ✅",7)
header(ws,4,["주차","담당(제안)","발표 주제","발표일","발표자 이름","발표자료 링크","상태"])
# 주차별 발표 슬롯 분배 (현장 15명이 4주에 1회씩 + 일부 ICU 녹화)
assign=[
 ("1주차","N06","심실세동/심정지(VF)"),("1주차","N11","심실빈맥(VT)"),("1주차","E01","아나필락시스"),("1주차","N01(파견)","완전방실차단 ※녹화 가능"),
 ("2주차","N07","급성 심근경색(STEMI)"),("2주차","N12","급성 폐색전증(PE)"),("2주차","E02","천식·COPD 악화"),("2주차","N02(파견)","급성 심부전·폐부종 ※녹화 가능"),
 ("3주차","N08","허혈성 뇌졸중"),("3주차","N13","저혈당"),("3주차","E04","뇌전증 지속상태"),("3주차","N03(파견)","출혈성 뇌졸중 ※녹화 가능"),
 ("4주차","N09","패혈증·패혈성 쇼크"),("4주차","N14","출혈성 쇼크·대량수혈"),("4주차","E05","약물 과용량/중독"),("4주차","N04(파견)","고칼륨혈증 ※녹화 가능"),
 ("예비","N10 / N15 / E03 / N05(파견)","예비 발표자·재발표·보강 주제(택1)"),
]
dv=DataValidation(type="list",formula1='"☐ 예정,▶ 진행,✅ 완료"',allow_blank=True); ws.add_data_validation(dv)
r=5
for item in assign:
    if len(item)==3 and item[0]!="예비":
        wk,who,topic=item
        cell(ws,r,1,wk,bold=True,fill=wkfill.get(wk,HDR),al=AC,color="FFFFFF")
        cell(ws,r,2,who,al=AC); cell(ws,r,3,topic)
        cell(ws,r,4,"",fill="FFFDE7",al=AC); cell(ws,r,5,"",fill="FFFDE7",al=AC)
        cell(ws,r,6,"",al=AC); st=cell(ws,r,7,"☐ 예정",al=AC); dv.add(st)
    else:
        wk,who,topic=item
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=1)
        cell(ws,r,1,"예비",bold=True,fill="455A64",al=AC,color="FFFFFF")
        cell(ws,r,2,who,al=AC); cell(ws,r,3,topic)
        cell(ws,r,4,"",fill="FFFDE7",al=AC); cell(ws,r,5,"",fill="FFFDE7",al=AC)
        cell(ws,r,6,"",al=AC); st=cell(ws,r,7,"☐ 예정",al=AC); dv.add(st)
    ws.row_dimensions[r].height=26; r+=1

# ════════════════════════════════════════════════════════════
# 5) 👥 팀 시나리오 조
# ════════════════════════════════════════════════════════════
ws=wb.create_sheet("👥 팀 시나리오 조")
for col,w in zip("ABCDEF",[8,10,34,22,30,12]): ws.column_dimensions[col].width=w
title_row(ws,"👥 팀 시나리오 조 편성 (혼성: 간호사+응급구조사)","조별로 매주 그 주차 시나리오 롤플레이 · 역할 분담 · 끝나면 상태 ✅",6)
header(ws,4,["조","근무","구성원(번호→이름 입력)","역할 분담","담당 시나리오(4주)","상태"])
teams=[
 ("A조","데이","N06, N07, E01","리더 / 기도·호흡 / 구조사 처치","주차별 해당 시나리오"),
 ("B조","데이","N08, N09, N10, E02, E03","리더 / 약물 / 기록 / 구조사 / 보조","주차별 해당 시나리오"),
 ("C조","이브닝","N11, N12, E04","리더 / 기도·호흡 / 구조사 처치","주차별 해당 시나리오"),
 ("D조","이브닝","N13, N14, N15, E05","리더 / 약물 / 기록 / 구조사","주차별 해당 시나리오"),
 ("ICU","파견","N01~N05","자가학습+퀴즈 / 가능일 합류","복귀·가능일 시나리오 참여"),
]
dv2=DataValidation(type="list",formula1='"☐ 예정,▶ 진행,✅ 완료"',allow_blank=True); ws.add_data_validation(dv2)
tcol={"A조":RED,"B조":ORANGE,"C조":BLUE,"D조":GREEN,"ICU":"455A64"}
r=5
for team,shift,mem,role,scen in teams:
    cell(ws,r,1,team,bold=True,fill=tcol[team],al=AC,color="FFFFFF")
    cell(ws,r,2,shift,al=AC); cell(ws,r,3,mem,fill="FFFDE7")
    cell(ws,r,4,role); cell(ws,r,5,scen)
    st=cell(ws,r,6,"☐ 예정",al=AC); dv2.add(st)
    ws.row_dimensions[r].height=40; r+=1
cell(ws,r+1,3,"※ 매주: 1주차 CPR / 2주차 STEMI / 3주차 뇌졸중·저혈당 / 4주차 패혈증 — 조마다 같은 주제를 자기 근무시간에 실습",al=AL,color="8B98A8")

# ════════════════════════════════════════════════════════════
# 6) ✅ 진행 체크보드
# ════════════════════════════════════════════════════════════
ws=wb.create_sheet("✅ 진행 체크보드")
for col,w in zip("ABCDEFG",[6,14,16,16,16,16,16]): ws.column_dimensions[col].width=w
title_row(ws,"✅ 진행 체크보드","각 칸 드롭다운(☐예정/▶진행/✅완료) · 전체 진행을 한눈에",7)
header(ws,4,["번호","이름(입력)","자가학습","개인 발표","팀 시나리오","퀴즈","비고"])
dv3=DataValidation(type="list",formula1='"☐ 예정,▶ 진행,✅ 완료"',allow_blank=True); ws.add_data_validation(dv3)
allp=[f"N{i:02d}" for i in range(1,16)]+[f"E{j:02d}" for j in range(1,6)]
r=5
for no in allp:
    fill=LGRAY if r%2 else "FFFFFF"
    cell(ws,r,1,no,bold=True,fill=fill,al=AC)
    cell(ws,r,2,"",fill="FFFDE7",al=AC)
    for c in range(3,7):
        x=cell(ws,r,c,"☐ 예정",fill=fill,al=AC); dv3.add(x)
    cell(ws,r,7,"파견" if no in allp[:5] else "",fill=fill,al=AC,color=RED)
    ws.row_dimensions[r].height=24; r+=1

wb.save(OUT)
print("저장:", OUT)
print("시트:", wb.sheetnames)

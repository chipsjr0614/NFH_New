#!/usr/bin/env python3
"""
statement-lint — 생성된 간호기록/진술문이 진술문 마스터(Nr.statement_*.xlsx)를
벗어나지 않는지 검증한다.

사용:
  python3 lint.py <검증대상 ...> [--master PATH] [--sheet 진술문] [--strict]

검증대상은 .json / .xlsx(수정본 포맷) / .html(BLOCKS 임베드) / .txt(줄당 진술문) 지원.

종료코드: 마스터를 벗어난 진술문(ERROR)이 하나라도 있으면 1, 아니면 0.
"""
import sys, os, re, json, glob, argparse, difflib

# ---------- 마스터 자동 탐색 ----------
def find_master(explicit=None):
    if explicit:
        return explicit
    here = os.getcwd()
    cands = []
    for pat in ("**/Nr.statement*.xlsx", "**/*진술문*마스터*.xlsx"):
        cands += glob.glob(os.path.join(here, pat), recursive=True)
    # 0_참고문서 / refs 우선
    cands.sort(key=lambda p: (0 if ("참고문서" in p or "/refs/" in p) else 1, len(p)))
    return cands[0] if cands else None

# ---------- 정규화 ----------
_NORM = re.compile(r"[\s,，·:：/、\.]+")
def norm(s):
    return _NORM.sub("", str(s).strip().lower())

# ---------- 토글 펼치기: "X 있음/없음" → ["X 있음","X 없음"] ----------
_TOG = ("있음", "없음", "강함", "약함")
def expand(stmt):
    s = str(stmt).strip()
    if "/" not in s:
        return [s]
    parts = [p.strip() for p in s.split("/") if p.strip()]
    m = re.match(r"^(.*?)(있음|없음|강함|약함)\s*$", parts[0])
    stem = m.group(1).strip() if m else ""
    out = []
    for p in parts:
        out.append((stem + " " + p) if (p in _TOG and stem) else p)
    return out or [s]

def is_value_item(s):
    return bool(re.search(r"[:：]\s*$", str(s)))

# ---------- 마스터 로드 ----------
def load_master(path, sheet):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    exact = set()
    for r in ws.iter_rows(values_only=True):
        for cell in r:
            if cell and isinstance(cell, str) and cell.strip():
                exact.add(cell.strip())
    normmap = {}
    for s in exact:
        normmap.setdefault(norm(s), s)
    return exact, normmap

# ---------- 대상에서 진술문 추출 ----------
def from_json(obj, out):
    if isinstance(obj, dict):
        for k, v in obj.items():
            if k in ("진술문", "간호 진술문", "간호진술문", "s") and isinstance(v, str):
                out.append(v)
            else:
                from_json(v, out)
    elif isinstance(obj, list):
        for x in obj:
            from_json(x, out)

def extract(path):
    ext = os.path.splitext(path)[1].lower()
    used = []
    if ext == ".json":
        from_json(json.load(open(path, encoding="utf-8")), used)
    elif ext == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header = [str(c).strip() if c else "" for c in rows[0]]
            col = next((i for i, h in enumerate(header) if "진술문" in h), None)
            if col is None:
                continue
            for r in rows[1:]:
                if col < len(r) and r[col] and str(r[col]).strip():
                    used.append(str(r[col]).strip())
    elif ext in (".html", ".htm", ".js"):
        txt = open(path, encoding="utf-8").read()
        m = re.search(r"BLOCKS\s*=\s*(\[.*?\]);", txt, re.S)
        if m:
            try:
                from_json(json.loads(m.group(1)), used)
            except Exception:
                pass
        if not used:  # fallback: "s":"..."
            used += re.findall(r'"s"\s*:\s*"([^"]+)"', txt)
    else:  # txt: 줄당 하나
        used += [ln.strip() for ln in open(path, encoding="utf-8") if ln.strip()]
    return used

# ---------- 검증 ----------
def lint_one(stmt, exact, normmap):
    """returns ('ok'|'variant'|'value'|'miss', detail)"""
    if is_value_item(stmt):
        return ("value", None)
    options = expand(stmt)
    results = []
    for opt in options:
        if opt in exact:
            results.append(("ok", None))
        elif norm(opt) in normmap:
            results.append(("variant", normmap[norm(opt)]))
        else:
            sug = difflib.get_close_matches(opt, list(exact), n=1, cutoff=0.6)
            results.append(("miss", sug[0] if sug else None))
    # 토글 전체에서 가장 나쁜 결과를 대표로
    order = {"miss": 3, "variant": 2, "value": 1, "ok": 0}
    return max(results, key=lambda r: order[r[0]])

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("targets", nargs="+")
    ap.add_argument("--master", default=None)
    ap.add_argument("--sheet", default="진술문")
    ap.add_argument("--strict", action="store_true", help="variant(띄어쓰기·문장부호 차이)도 실패로 처리")
    args = ap.parse_args()

    master = find_master(args.master)
    if not master or not os.path.exists(master):
        print("✗ 진술문 마스터를 찾지 못했습니다. --master 로 경로를 지정하세요.")
        sys.exit(2)
    print(f"📖 마스터: {os.path.relpath(master)}")
    exact, normmap = load_master(master, args.sheet)
    print(f"   진술문 {len(exact):,}개 로드\n")

    total_err = 0
    for tgt in args.targets:
        if not os.path.exists(tgt):
            print(f"⚠️  대상 없음: {tgt}"); continue
        used = extract(tgt)
        seen, misses, variants, values = set(), [], [], 0
        for s in used:
            if s in seen:
                continue
            seen.add(s)
            kind, detail = lint_one(s, exact, normmap)
            if kind == "miss":
                misses.append((s, detail))
            elif kind == "variant":
                variants.append((s, detail))
            elif kind == "value":
                values += 1

        print(f"── {os.path.relpath(tgt)} — 고유 진술문 {len(seen)}개")
        if misses:
            print(f"   ❌ 마스터에 없음 {len(misses)}건 (벗어남):")
            for s, sug in misses:
                tail = f"   → 가까운 후보: «{sug}»" if sug else "   → 후보 없음"
                print(f"      • «{s}»{tail}")
        if variants:
            print(f"   ⚠️  표기 변형 {len(variants)}건 (띄어쓰기·문장부호만 다름):")
            for s, orig in variants:
                print(f"      • «{s}»  ↔ 마스터 «{orig}»")
        if values:
            print(f"   ℹ️  값 입력형(플레이스홀더) {values}건 — 검증 제외")
        if not misses and not variants:
            print("   ✅ 전부 마스터 진술문과 일치")
        print()
        total_err += len(misses) + (len(variants) if args.strict else 0)

    if total_err:
        print(f"결과: ❌ 벗어난 진술문 {total_err}건 — 마스터에 맞게 수정 필요")
        sys.exit(1)
    print("결과: ✅ 통과 — 모든 진술문이 마스터 범위 안")
    sys.exit(0)

if __name__ == "__main__":
    main()
